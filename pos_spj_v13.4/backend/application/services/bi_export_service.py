"""BiExportService — exporta el dashboard ejecutivo a Excel / PDF / CSV.

Toma el DashboardPayload (dict) + metadatos (usuario, rango, sucursal, fecha de
generación) y produce un archivo. Librerías opcionales (openpyxl / reportlab)
con fallback a CSV. No consulta la DB: recibe datos ya calculados por el servicio.
"""
from __future__ import annotations

import csv
import logging
import os
from datetime import datetime

logger = logging.getLogger("spj.bi.export")


def _fmt(value, unit) -> str:
    try:
        value = float(value or 0)
    except Exception:
        return str(value)
    if unit == "%":
        return f"{value:.2f}%"
    if unit == "x":
        return f"{value:.2f}x"
    if unit == "":
        return f"{int(round(value)):,}"
    return f"${value:,.2f}"


class BiExportService:
    def export_summary(self, payload: dict, meta: dict, filepath: str,
                       fmt: str = "xlsx") -> str:
        """Exporta el resumen ejecutivo. Devuelve la ruta escrita."""
        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
        meta = self._complete_meta(meta)
        if fmt == "xlsx":
            return self._xlsx(payload, meta, filepath)
        if fmt == "pdf":
            return self._pdf(payload, meta, filepath)
        return self._csv(payload, meta, filepath)

    # ── helpers ───────────────────────────────────────────────────────────────

    @staticmethod
    def _complete_meta(meta: dict) -> dict:
        m = dict(meta or {})
        m.setdefault("generado", datetime.now().strftime("%Y-%m-%d %H:%M"))
        m.setdefault("usuario", "—")
        m.setdefault("rango", "—")
        m.setdefault("sucursal", "Todas")
        return m

    @staticmethod
    def _meta_rows(meta: dict) -> list[list]:
        return [
            ["Reporte BI — Resumen ejecutivo"],
            ["Generado", meta["generado"]],
            ["Usuario", meta["usuario"]],
            ["Periodo", meta["rango"]],
            ["Sucursal", meta["sucursal"]],
            [],
        ]

    @staticmethod
    def _kpi_rows(payload: dict) -> list[list]:
        rows = [["KPI", "Valor", "vs anterior"]]
        for k in payload.get("kpis", []):
            dp = k.get("delta_pct")
            pts = k.get("delta_points")
            comp = (f"{dp:+.1f}%" if dp is not None
                    else f"{pts:+.2f} pp" if pts is not None else "—")
            rows.append([k.get("title", ""), _fmt(k.get("value"), k.get("unit", "")), comp])
        return rows

    def _csv(self, payload, meta, filepath) -> str:
        if not filepath.endswith(".csv"):
            filepath = os.path.splitext(filepath)[0] + ".csv"
        with open(filepath, "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.writer(fh)
            for r in self._meta_rows(meta):
                w.writerow(r)
            for r in self._kpi_rows(payload):
                w.writerow(r)
            w.writerow([])
            w.writerow(["Alertas"])
            for a in payload.get("alerts", []):
                w.writerow([a.get("title", ""), a.get("detail", "")])
            w.writerow([])
            w.writerow(["Insights"])
            for i in payload.get("insights", []):
                w.writerow([i.get("title", ""), i.get("detail", "")])
        return filepath

    def _xlsx(self, payload, meta, filepath) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            logger.warning("openpyxl ausente — exportando CSV")
            return self._csv(payload, meta, filepath)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resumen"
        for r in self._meta_rows(meta):
            ws.append(r)
        ws["A1"].font = Font(bold=True, size=14)
        header_at = ws.max_row + 1
        for r in self._kpi_rows(payload):
            ws.append(r)
        for cell in ws[header_at]:
            cell.font = Font(bold=True)
        # Highlights
        hl = payload.get("highlights", {})
        if hl:
            ws.append([])
            ws.append(["Destacados"])
            ws[ws.max_row][0].font = Font(bold=True)
            for h in hl.values():
                ws.append([h.get("title", ""), h.get("name", ""),
                           _fmt(h.get("value"), h.get("unit", "$")),
                           f"{h.get('share_pct', 0):.1f}%"])
        wa = wb.create_sheet("Alertas")
        wa.append(["Nivel", "Alerta", "Detalle"])
        for a in payload.get("alerts", []):
            wa.append([a.get("level", ""), a.get("title", ""), a.get("detail", "")])
        wi = wb.create_sheet("Insights")
        wi.append(["Insight", "Detalle"])
        for i in payload.get("insights", []):
            wi.append([i.get("title", ""), i.get("detail", "")])
        if not filepath.endswith(".xlsx"):
            filepath = os.path.splitext(filepath)[0] + ".xlsx"
        wb.save(filepath)
        return filepath

    def _pdf(self, payload, meta, filepath) -> str:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import (
                Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle)
        except ImportError:
            logger.warning("reportlab ausente — exportando CSV")
            return self._csv(payload, meta, filepath)
        if not filepath.endswith(".pdf"):
            filepath = os.path.splitext(filepath)[0] + ".pdf"
        doc = SimpleDocTemplate(filepath, pagesize=letter)
        styles = getSampleStyleSheet()
        el = [Paragraph("<b>Reporte BI — Resumen ejecutivo</b>", styles["Title"])]
        el.append(Paragraph(
            f"Generado: {meta['generado']} · Usuario: {meta['usuario']} · "
            f"Periodo: {meta['rango']} · Sucursal: {meta['sucursal']}", styles["Normal"]))
        el.append(Spacer(1, 12))
        tbl = Table(self._kpi_rows(payload), repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ]))
        el.append(tbl)
        alerts = payload.get("alerts", [])
        if alerts:
            el.append(Spacer(1, 12))
            el.append(Paragraph("<b>Alertas</b>", styles["Heading3"]))
            for a in alerts:
                el.append(Paragraph(f"• {a.get('title','')}: {a.get('detail','')}",
                                    styles["Normal"]))
        insights = payload.get("insights", [])
        if insights:
            el.append(Spacer(1, 8))
            el.append(Paragraph("<b>Insights</b>", styles["Heading3"]))
            for i in insights:
                el.append(Paragraph(f"• {i.get('title','')}: {i.get('detail','')}",
                                    styles["Normal"]))
        doc.build(el)
        return filepath
