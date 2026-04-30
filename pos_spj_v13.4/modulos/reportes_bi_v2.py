# -*- coding: utf-8 -*-
# modulos/reportes_bi_v2.py
from modulos.design_tokens import Colors, Spacing, Typography, Borders, Shadows
from modulos.ui_components import (
    create_primary_button, create_success_button, create_danger_button,
    create_secondary_button, create_input, create_combo, create_card,
    create_heading, create_subheading, create_caption, apply_tooltip,
    FilterBar, EmptyStateWidget, LoadingIndicator, DataTableWithFilters, confirm_action,
    PageHeader, Toast,
)
from modulos.spj_styles import spj_btn, apply_btn_styles
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QLineEdit,
    QComboBox, QMessageBox, QFormLayout, QTableWidget, QTableWidgetItem,
    QHeaderView, QGridLayout, QGroupBox, QFrame, QSplitter, QTabWidget,
    QAbstractItemView, QDialog, QCheckBox, QListWidget, QListWidgetItem,
    QSizePolicy, QAction, QMenu, QToolBar, QStatusBar, QProgressBar,
    QCompleter, QDateEdit, QSpinBox, QDoubleSpinBox,
    QFileDialog
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont

class ModuloReportesBIv2(QWidget):
    """
    Dashboard Corporativo de Business Intelligence.
    Cero SQL. Toda la data proviene del BIService.
    """
    def __init__(self, container, parent=None):
        super().__init__(parent)
        self.container = container
        self.sucursal_id = 1
        self._last_data = {}
        self.init_ui()
        self._wire_business_events()

    def set_usuario_actual(self, usuario: str, rol: str = "cajero") -> None:
        """Recibe el usuario activo al cambiar de sesión."""
        self.usuario_actual = usuario
        self.rol_actual = rol

    def set_sucursal(self, sucursal_id: int, nombre_sucursal: str = ""):
        """Recibe la sucursal activa y refresca dashboard."""
        self.sucursal_id = sucursal_id
        self.cargar_dashboard()

    def init_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        layout_principal = QVBoxLayout()
        layout_principal.setSpacing(10)
        outer.addLayout(layout_principal)

        # --- HEADER (PageHeader: título + subtítulo + acciones) ---
        self.page_header = PageHeader(
            self,
            title="📈 Inteligencia Comercial",
            subtitle="Dashboard ejecutivo de BI",
        )

        # Período
        self.cmb_rango = create_combo(self, ["Hoy", "Esta Semana", "Este Mes"])
        self.cmb_rango.currentTextChanged.connect(self.cargar_dashboard)
        self.page_header.add_action(QLabel("Período:"))
        self.page_header.add_action(self.cmb_rango)

        # Refrescar
        self.btn_actualizar = create_secondary_button(self, "🔄 Refrescar", "Actualizar datos del dashboard")
        self.btn_actualizar.clicked.connect(self.cargar_dashboard)
        self.page_header.add_action(self.btn_actualizar)

        # Exportar
        btn_excel = create_success_button(self, "📊 Excel", "Exportar dashboard a Excel (.xlsx)")
        btn_excel.clicked.connect(lambda: self._exportar("excel"))
        self.page_header.add_action(btn_excel)

        btn_pdf = create_danger_button(self, "📄 PDF", "Exportar dashboard a PDF")
        btn_pdf.clicked.connect(lambda: self._exportar("pdf"))
        self.page_header.add_action(btn_pdf)

        layout_principal.addWidget(self.page_header)

        self.filter_bar = FilterBar(
            self,
            placeholder="Buscar producto, cliente o cajero…",
            combo_filters={"vista": ["Resumen", "Rankings", "Rentabilidad", "Cajeros"]}
        )
        self.filter_bar.filters_changed.connect(self._on_global_filters_changed)
        layout_principal.addWidget(self.filter_bar)

        self.loading_dashboard = LoadingIndicator("Actualizando dashboard BI…", self)
        self.loading_dashboard.hide()
        layout_principal.addWidget(self.loading_dashboard)

        self.tabs_bi = QTabWidget()
        self.tabs_bi.setDocumentMode(True)
        layout_principal.addWidget(self.tabs_bi)

        self._build_tab_resumen()
        self._build_tab_visual_dashboard()
        self._build_tab_rankings()
        self._build_tab_rentabilidad()
        self._build_tab_cajeros()
        self._build_tab_forecast()
        self._build_tab_decision_engine()
        self._build_tab_franchise()

    def _crear_tab_contenedor(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)
        return tab, layout

    def _build_tab_resumen(self):
        tab, layout = self._crear_tab_contenedor()

        kpi_layout = QHBoxLayout()
        self.lbl_kpi_ingresos = self._crear_tarjeta_kpi("💵 Ingresos Totales", "$0.00")
        self.lbl_kpi_ticket = self._crear_tarjeta_kpi("🧾 Ticket Promedio", "$0.00")
        self.lbl_kpi_ventas = self._crear_tarjeta_kpi("🛒 Num. Ventas", "0")
        self.lbl_kpi_clientes = self._crear_tarjeta_kpi("👥 Clientes Únicos", "0")
        kpi_layout.addWidget(self.lbl_kpi_ingresos)
        kpi_layout.addWidget(self.lbl_kpi_ticket)
        kpi_layout.addWidget(self.lbl_kpi_ventas)
        kpi_layout.addWidget(self.lbl_kpi_clientes)
        layout.addLayout(kpi_layout)

        self.lbl_comparativa = create_caption(self, "")
        self.lbl_comparativa.setAlignment(Qt.AlignCenter)
        self.lbl_comparativa.hide()
        layout.addWidget(self.lbl_comparativa)

        self._lbl_resumen_alertas = QLabel("Alertas: esperando actualización de datos.")
        self._lbl_resumen_alertas.setStyleSheet("color:#6c757d;")
        layout.addWidget(self._lbl_resumen_alertas)

        accesos = QGroupBox("Acceso rápido")
        accesos_lay = QHBoxLayout(accesos)
        for idx, texto in [
            (2, "Ventas / Rankings"),
            (3, "Rentabilidad"),
            (4, "Cajeros"),
            (5, "Forecast"),
            (6, "Sugerencias"),
            (7, "Sucursales"),
        ]:
            btn = create_secondary_button(self, texto, f"Ir a {texto}")
            btn.clicked.connect(lambda _, i=idx: self.tabs_bi.setCurrentIndex(i))
            accesos_lay.addWidget(btn)
        layout.addWidget(accesos)
        layout.addStretch()
        self.tabs_bi.addTab(tab, "Resumen Ejecutivo")

    def _build_tab_visual_dashboard(self):
        """Dashboard visual moderno con QWebEngineView + Apache ECharts."""
        tab, layout = self._crear_tab_contenedor()

        self._chart_empty = EmptyStateWidget(
            "Sin datos para graficar",
            "No hay movimientos suficientes para construir el dashboard visual.",
            "📉",
            self,
        )
        self._chart_empty.hide()

        try:
            from PyQt5.QtWebEngineWidgets import QWebEngineView
            self._chart_view = QWebEngineView(self)
            layout.addWidget(self._chart_view, 1)
            layout.addWidget(self._chart_empty)
        except Exception:
            self._chart_view = None
            fallback = QLabel("QWebEngine no disponible en este entorno.\nSe mostrará solo KPI tabular.")
            fallback.setStyleSheet("color:#6c757d; padding:12px;")
            fallback.setAlignment(Qt.AlignCenter)
            layout.addWidget(fallback)
        self.tabs_bi.addTab(tab, "Dashboard Visual")

    def _build_tab_rankings(self):
        tab, layout = self._crear_tab_contenedor()
        toolbar = QHBoxLayout()
        btn_ref_rank = create_primary_button(self, "🔄 Actualizar", "Actualizar rankings")
        btn_ref_rank.clicked.connect(self.cargar_dashboard)
        btn_export_rank = create_secondary_button(self, "📊 Exportar", "Exportar rankings a Excel")
        btn_export_rank.clicked.connect(lambda: self._exportar("excel"))
        self._lbl_rankings_estado = QLabel("Sin datos cargados.")
        self._lbl_rankings_estado.setStyleSheet("color:#6c757d; font-size:11px;")
        toolbar.addWidget(btn_ref_rank)
        toolbar.addWidget(btn_export_rank)
        toolbar.addStretch()
        toolbar.addWidget(self._lbl_rankings_estado)
        layout.addLayout(toolbar)

        self._tabs_rankings = QTabWidget()
        self._tabs_rankings.setDocumentMode(True)
        self._grp_top = self._crear_tabla_ranking("⭐ Productos Más Vendidos", ["Producto", "Cant.", "Ingresos"])
        self.tabla_top = self._grp_top._tabla.table
        self._grp_lentos = self._crear_tabla_ranking("🐢 Productos Lentos", ["Producto", "Cant.", "Ingresos"])
        self.tabla_lentos = self._grp_lentos._tabla.table
        self._grp_vips = self._crear_tabla_ranking("👑 Clientes Recurrentes (VIP)", ["Cliente", "Visitas", "Gasto Total"])
        self.tabla_vips = self._grp_vips._tabla.table
        self._tabs_rankings.addTab(self._grp_top, "Más vendidos")
        self._tabs_rankings.addTab(self._grp_lentos, "Lentos")
        self._tabs_rankings.addTab(self._grp_vips, "Clientes VIP")
        layout.addWidget(self._tabs_rankings)
        self.tabs_bi.addTab(tab, "Ventas / Rankings")

    def _build_tab_rentabilidad(self):
        tab, layout = self._crear_tab_contenedor()
        self._build_rentabilidad_section(layout)
        self.tabs_bi.addTab(tab, "Rentabilidad")

    def _build_tab_cajeros(self):
        tab, layout = self._crear_tab_contenedor()
        self._build_cajeros_section(layout)
        self.tabs_bi.addTab(tab, "Cajeros")

    def _build_tab_forecast(self):
        tab, layout = self._crear_tab_contenedor()
        self._build_forecast_section(layout)
        self.tabs_bi.addTab(tab, "Forecast / Planeación")

    def _build_tab_decision_engine(self):
        tab, layout = self._crear_tab_contenedor()
        self._build_decision_engine_section(layout)
        self.tabs_bi.addTab(tab, "Sugerencias")

    def _build_tab_franchise(self):
        tab, layout = self._crear_tab_contenedor()
        self._build_franchise_section(layout)
        self.tabs_bi.addTab(tab, "Sucursales / Franquicias")

    def _build_rentabilidad_section(self, parent_layout):
        """Tabla de rentabilidad por producto: margen, rotación, contribución."""
        from PyQt5.QtWidgets import QGroupBox
        grp = QGroupBox("💰 Rentabilidad por Producto (Margen Bruto)")
        grp.setObjectName("styledGroup")
        lay = QVBoxLayout(grp)

        toolbar = QHBoxLayout()
        btn_rent = create_primary_button(self, "📊 Calcular Rentabilidad", "Calcular rentabilidad por producto")
        btn_rent.clicked.connect(self._cargar_rentabilidad)
        btn_export = create_secondary_button(self, "📥 Exportar CSV", "Exportar reporte a CSV")
        btn_export.clicked.connect(self._exportar_rentabilidad_csv)
        toolbar.addWidget(btn_rent); toolbar.addWidget(btn_export); toolbar.addStretch()
        lay.addLayout(toolbar)

        self._tbl_rent = QTableWidget()
        self._tbl_rent.setColumnCount(8)
        self._tbl_rent.setHorizontalHeaderLabels([
            "Producto", "Categoría", "Unidades", "Ingresos",
            "Costo Total", "Margen $", "Margen %", "ABC"
        ])
        self._tbl_rent.setEditTriggers(QTableWidget.NoEditTriggers)
        self._tbl_rent.setSelectionBehavior(QTableWidget.SelectRows)
        self._tbl_rent.setAlternatingRowColors(True)
        self._tbl_rent.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self._tbl_rent.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        hh = self._tbl_rent.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.Stretch)
        self._tbl_rent.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        lay.addWidget(self._tbl_rent)
        self._lbl_rent_estado = QLabel("Calculando rentabilidad…")
        self._lbl_rent_estado.setStyleSheet("color:#6c757d; font-size:11px;")
        lay.addWidget(self._lbl_rent_estado)

        parent_layout.addWidget(grp)
        self._cargar_rentabilidad()

    # ── Rendimiento Cajeros: ranking por frecuencia y volumen ─────────────────

    def _build_cajeros_section(self, parent_layout):
        """Ranking de cajeros por # transacciones, volumen y ticket prom. Fase 2."""
        grp = QGroupBox("📊 Rendimiento de Cajeros")
        grp.setStyleSheet(
            "QGroupBox { font-weight:bold; border:1px solid #dee2e6;"
            " border-radius:6px; margin-top:10px; padding-top:8px; }"
        )
        lay = QVBoxLayout(grp)

        toolbar = QHBoxLayout()
        btn_caj = create_primary_button(self, "📋 Cargar Ranking Cajeros",
                                        "Calcular ranking de cajeros por transacciones y volumen")
        btn_caj.clicked.connect(self._cargar_cajeros)
        toolbar.addWidget(btn_caj)
        toolbar.addStretch()
        self._lbl_caj_estado = QLabel("Haz clic para ver el rendimiento de cajeros del mes.")
        self._lbl_caj_estado.setStyleSheet("color:#888; font-size:11px;")
        toolbar.addWidget(self._lbl_caj_estado)
        lay.addLayout(toolbar)

        self._tbl_caj = QTableWidget()
        self._tbl_caj.setColumnCount(6)
        self._tbl_caj.setHorizontalHeaderLabels([
            "Cajero", "# Ventas", "Total $", "Ticket Prom $", "Descuentos $", "Días Activo"
        ])
        hh = self._tbl_caj.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.Stretch)
        self._tbl_caj.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._tbl_caj.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._tbl_caj.setAlternatingRowColors(True)
        self._tbl_caj.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self._tbl_caj.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self._tbl_caj.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        lay.addWidget(self._tbl_caj)

        parent_layout.addWidget(grp)

    def _cargar_cajeros(self):
        """Carga ranking de cajeros via AnalyticsEngine.get_ranking_cajeros() (BI unificado)."""
        self._tbl_caj.setRowCount(0)
        try:
            # BI unificado: fuente única analytics_engine
            analytics = getattr(self.container, 'analytics_engine', None)
            if analytics is None:
                self._lbl_caj_estado.setText("AnalyticsEngine no disponible.")
                return
            rango = self.cmb_rango.currentText()
            rango_key = "hoy" if "hoy" in rango.lower() else \
                        "semana" if "semana" in rango.lower() else "mes"
            # Calcular fechas para el ranking
            from datetime import datetime, timedelta
            hoy = datetime.now()
            if rango_key == 'hoy':
                fecha_inicio = fecha_fin = hoy.strftime('%Y-%m-%d')
            elif rango_key == 'semana':
                fecha_inicio = (hoy - timedelta(days=hoy.weekday())).strftime('%Y-%m-%d')
                fecha_fin = hoy.strftime('%Y-%m-%d')
            else:  # mes
                fecha_inicio = hoy.replace(day=1).strftime('%Y-%m-%d')
                fecha_fin = hoy.strftime('%Y-%m-%d')
            rows = analytics.get_ranking_cajeros(self.sucursal_id, fecha_inicio, fecha_fin, limite=20)
            self._lbl_caj_estado.setText(f"{len(rows)} cajeros encontrados.")
            for i, r in enumerate(rows):
                self._tbl_caj.insertRow(i)
                vals = [
                    str(r.get("cajero", "(sin usuario)")),
                    str(r.get("num_ventas", 0)),
                    f"${float(r.get('total_ventas', 0)):,.2f}",
                    f"${float(r.get('ticket_promedio', 0)):,.2f}",
                    f"${float(r.get('total_descuentos', 0)):,.2f}",
                    str(r.get("dias_activo", 0)),
                ]
                for j, v in enumerate(vals):
                    item = QTableWidgetItem(v)
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    if i == 0:  # Top cajero destacado
                        item.setForeground(
                            __import__('PyQt5.QtGui', fromlist=['QColor']).QColor('#27ae60'))
                    self._tbl_caj.setItem(i, j, item)
        except Exception as exc:
            self._lbl_caj_estado.setText(f"Error: {exc}")

    # ── ActionableForecast: plan de compras y riesgos ─────────────────────────

    def _build_forecast_section(self, parent_layout):
        """Plan de compras semanal + análisis de riesgos de inventario. Fase 5."""
        grp = QGroupBox("🔮 Forecast & Abastecimiento (ActionableForecast)")
        grp.setStyleSheet(
            "QGroupBox { font-weight:bold; border:1px solid #dee2e6;"
            " border-radius:6px; margin-top:10px; padding-top:8px; }"
        )
        lay = QVBoxLayout(grp)

        toolbar = QHBoxLayout()
        btn_compras = create_primary_button(self, "🛒 Plan Compras Semanal",
                                            "Generar plan de compras basado en demanda histórica")
        btn_compras.clicked.connect(self._cargar_plan_compras)
        btn_riesgos = create_secondary_button(self, "⚠️ Análisis de Riesgos",
                                              "Detectar productos con riesgo de desabasto")
        btn_riesgos.clicked.connect(self._cargar_riesgos_inventario)
        toolbar.addWidget(btn_compras)
        toolbar.addWidget(btn_riesgos)
        toolbar.addStretch()
        self._lbl_fc_estado = QLabel("Selecciona una acción para generar el forecast.")
        self._lbl_fc_estado.setStyleSheet("color:#888; font-size:11px;")
        toolbar.addWidget(self._lbl_fc_estado)
        lay.addLayout(toolbar)

        self._tbl_fc = QTableWidget()
        self._tbl_fc.setColumnCount(6)
        self._tbl_fc.setHorizontalHeaderLabels([
            "Producto", "Stock Actual", "Demanda/día", "Días Stock",
            "Comprar", "Costo Estimado $"
        ])
        hh = self._tbl_fc.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.Stretch)
        self._tbl_fc.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._tbl_fc.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._tbl_fc.setAlternatingRowColors(True)
        self._tbl_fc.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self._tbl_fc.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self._tbl_fc.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        lay.addWidget(self._tbl_fc)

        parent_layout.addWidget(grp)

    def _cargar_plan_compras(self):
        """Carga el plan de compras via ActionableForecastService.plan_compras_semanal()."""
        self._tbl_fc.setRowCount(0)
        self._tbl_fc.setHorizontalHeaderLabels([
            "Producto", "Stock Actual", "Demanda/día", "Días Stock",
            "Comprar", "Costo Estimado $"
        ])
        try:
            svc = getattr(self.container, "actionable_forecast", None)
            if svc is None:
                self._lbl_fc_estado.setText("ActionableForecastService no disponible.")
                return
            rows = svc.plan_compras_semanal(sucursal_id=self.sucursal_id)
            self._lbl_fc_estado.setText(f"{len(rows)} productos en plan de compras.")
            for i, r in enumerate(rows):
                self._tbl_fc.insertRow(i)
                prioridad = r.get("prioridad", "")
                vals = [
                    str(r.get("producto", "")),
                    f"{float(r.get('stock_actual', 0)):.2f}",
                    f"{float(r.get('demanda_diaria', 0)):.2f}",
                    f"{float(r.get('dias_stock', 0)):.1f}",
                    f"{float(r.get('comprar_kg', r.get('cantidad_comprar', 0))):.2f}",
                    f"${float(r.get('costo_est', r.get('costo_estimado', 0))):,.2f}",
                ]
                for j, v in enumerate(vals):
                    item = QTableWidgetItem(v)
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    if prioridad == "alta" and j == 0:
                        item.setForeground(
                            __import__('PyQt5.QtGui', fromlist=['QColor']).QColor('#e74c3c'))
                    self._tbl_fc.setItem(i, j, item)
        except Exception as exc:
            self._lbl_fc_estado.setText(f"Error: {exc}")

    def _cargar_riesgos_inventario(self):
        """Carga riesgos de inventario via ActionableForecastService.analisis_riesgos()."""
        self._tbl_fc.setRowCount(0)
        self._tbl_fc.setHorizontalHeaderLabels([
            "Tipo Riesgo", "Producto", "Días Stock", "Stock Actual",
            "Prioridad", "Acción Sugerida"
        ])
        try:
            svc = getattr(self.container, "actionable_forecast", None)
            if svc is None:
                self._lbl_fc_estado.setText("ActionableForecastService no disponible.")
                return
            rows = svc.analisis_riesgos(sucursal_id=self.sucursal_id)
            self._lbl_fc_estado.setText(f"{len(rows)} riesgos detectados.")
            for i, r in enumerate(rows):
                self._tbl_fc.insertRow(i)
                vals = [
                    str(r.get("tipo", "")),
                    str(r.get("producto", "")),
                    f"{float(r.get('dias_stock', 0)):.1f}",
                    f"{float(r.get('stock_actual', 0)):.2f}",
                    str(r.get("prioridad", "")),
                    str(r.get("accion", r.get("accion_sugerida", "")))[:60],
                ]
                for j, v in enumerate(vals):
                    item = QTableWidgetItem(v)
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    if r.get("prioridad") == "crítica":
                        item.setForeground(
                            __import__('PyQt5.QtGui', fromlist=['QColor']).QColor('#e74c3c'))
                    self._tbl_fc.setItem(i, j, item)
        except Exception as exc:
            self._lbl_fc_estado.setText(f"Error: {exc}")

    # ── DecisionEngine: sugerencias accionables ───────────────────────────────

    def _build_decision_engine_section(self, parent_layout):
        """Panel de sugerencias del DecisionEngine (FASE 5 — solo lectura)."""
        grp = QGroupBox("🤖 Sugerencias Accionables (DecisionEngine)")
        grp.setStyleSheet(
            "QGroupBox { font-weight:bold; border:1px solid #dee2e6;"
            " border-radius:6px; margin-top:10px; padding-top:8px; }"
        )
        lay = QVBoxLayout(grp)

        toolbar = QHBoxLayout()
        btn_gen = create_primary_button(self, "🔍 Generar Sugerencias",
                                        "Analizar datos y generar sugerencias accionables")
        btn_gen.clicked.connect(self._cargar_decision_engine)
        toolbar.addWidget(btn_gen)
        toolbar.addStretch()
        self._lbl_de_estado = QLabel("Haz clic en 'Generar Sugerencias' para analizar.")
        self._lbl_de_estado.setStyleSheet("color:#888; font-size:11px;")
        toolbar.addWidget(self._lbl_de_estado)
        lay.addLayout(toolbar)

        self._tbl_de = QTableWidget()
        self._tbl_de.setColumnCount(5)
        self._tbl_de.setHorizontalHeaderLabels([
            "Prioridad", "Tipo", "Sugerencia", "Impacto Estimado", "Acción Propuesta"
        ])
        hh = self._tbl_de.horizontalHeader()
        hh.setSectionResizeMode(2, QHeaderView.Stretch)
        hh.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self._tbl_de.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._tbl_de.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._tbl_de.setAlternatingRowColors(True)
        self._tbl_de.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self._tbl_de.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self._tbl_de.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        lay.addWidget(self._tbl_de)

        parent_layout.addWidget(grp)

    def _cargar_decision_engine(self):
        """Invoca DecisionEngine.generar_sugerencias() y muestra los resultados."""
        self._tbl_de.setRowCount(0)
        try:
            engine = getattr(self.container, "decision_engine", None)
            if engine is None:
                self._lbl_de_estado.setText("DecisionEngine no disponible en este contenedor.")
                return
            sugs = engine.generar_sugerencias(sucursal_id=self.sucursal_id)
            self._lbl_de_estado.setText(
                f"{len(sugs)} sugerencias generadas — solo lectura, no se ejecuta nada."
            )
            for i, s in enumerate(sugs):
                self._tbl_de.insertRow(i)
                accion = s.get("accion_propuesta", {})
                accion_txt = accion.get("descripcion", str(accion)) if isinstance(accion, dict) else str(accion)
                vals = [
                    s.get("prioridad", ""),
                    s.get("tipo", ""),
                    s.get("titulo", "") + (" — " + s.get("detalle", "") if s.get("detalle") else ""),
                    s.get("impacto_estimado", ""),
                    accion_txt[:80],
                ]
                for j, v in enumerate(vals):
                    item = QTableWidgetItem(str(v))
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    self._tbl_de.setItem(i, j, item)
        except Exception as exc:
            self._lbl_de_estado.setText(f"Error: {exc}")

    # ── FranchiseManager: ranking multi-sucursal ──────────────────────────────

    def _build_franchise_section(self, parent_layout):
        """Ranking de sucursales via FranchiseManager (Fase 6 — multi-franquicia)."""
        grp = QGroupBox("🏪 Ranking de Sucursales (FranchiseManager)")
        grp.setStyleSheet(
            "QGroupBox { font-weight:bold; border:1px solid #dee2e6;"
            " border-radius:6px; margin-top:10px; padding-top:8px; }"
        )
        lay = QVBoxLayout(grp)

        toolbar = QHBoxLayout()
        btn_rank = create_primary_button(self, "🏆 Calcular Ranking",
                                         "Calcular ranking de sucursales por ventas y rentabilidad")
        btn_rank.clicked.connect(self._cargar_franchise_ranking)
        toolbar.addWidget(btn_rank)
        toolbar.addStretch()
        self._lbl_fm_estado = QLabel("Haz clic en 'Calcular Ranking' para comparar sucursales.")
        self._lbl_fm_estado.setStyleSheet("color:#888; font-size:11px;")
        toolbar.addWidget(self._lbl_fm_estado)
        lay.addLayout(toolbar)

        self._tbl_fm = QTableWidget()
        self._tbl_fm.setColumnCount(6)
        self._tbl_fm.setHorizontalHeaderLabels([
            "Sucursal", "Ventas $", "# Transacciones", "Ticket Prom $",
            "Margen Bruto %", "Rank"
        ])
        hh = self._tbl_fm.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.Stretch)
        self._tbl_fm.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._tbl_fm.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._tbl_fm.setAlternatingRowColors(True)
        self._tbl_fm.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self._tbl_fm.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self._tbl_fm.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        lay.addWidget(self._tbl_fm)

        parent_layout.addWidget(grp)

    def _cargar_franchise_ranking(self):
        """Carga ranking de sucursales via FranchiseManager."""
        self._tbl_fm.setRowCount(0)
        try:
            fm = getattr(self.container, "franchise_manager", None)
            if fm is None:
                self._lbl_fm_estado.setText("FranchiseManager no disponible.")
                return
            rows = fm.ranking_sucursales()
            self._lbl_fm_estado.setText(f"{len(rows)} sucursales analizadas.")
            for i, r in enumerate(rows):
                self._tbl_fm.insertRow(i)
                vals = [
                    str(r.get("nombre", r.get("sucursal_id", ""))),
                    f"${float(r.get('total_ventas', 0)):,.2f}",
                    str(r.get("num_transacciones", 0)),
                    f"${float(r.get('ticket_promedio', 0)):,.2f}",
                    f"{float(r.get('margen_bruto_pct', 0)):.1f}%",
                    str(r.get("rank", i + 1)),
                ]
                for j, v in enumerate(vals):
                    item = QTableWidgetItem(str(v))
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    self._tbl_fm.setItem(i, j, item)
        except Exception as exc:
            self._lbl_fm_estado.setText(f"Error: {exc}")

    def _cargar_rentabilidad(self):
        """Carga el reporte de rentabilidad por producto para el período activo (BI unificado)."""
        self._tbl_rent.setRowCount(0)
        rango = self.cmb_rango.currentText()
        
        # Calcular fechas según el rango seleccionado
        from datetime import datetime, timedelta
        hoy = datetime.now()
        if "hoy" in rango.lower():
            fecha_inicio = fecha_fin = hoy.strftime('%Y-%m-%d')
        elif "semana" in rango.lower():
            fecha_inicio = (hoy - timedelta(days=hoy.weekday())).strftime('%Y-%m-%d')
            fecha_fin = hoy.strftime('%Y-%m-%d')
        else:  # mes
            fecha_inicio = hoy.replace(day=1).strftime('%Y-%m-%d')
            fecha_fin = hoy.strftime('%Y-%m-%d')
        
        try:
            # ✅ FASE 2: Usar AnalyticsEngine para rentabilidad de productos
            analytics = getattr(self.container, 'analytics_engine', None)
            if not analytics:
                raise RuntimeError("AnalyticsEngine no disponible")
            
            # Usar método del servicio unificado
            rows_raw = analytics.product_profitability(fecha_inicio, fecha_fin, self.sucursal_id, limit=50)
            # Convertir al formato esperado por la UI
            rows = []
            for r in rows_raw:
                # Necesitamos obtener nombre y categoría del producto
                prod_info = self._get_producto_info(r['producto_id'])
                rows.append((
                    prod_info.get('nombre', f"Prod {r['producto_id']}"),
                    prod_info.get('categoria', ''),
                    0,  # unidades (no disponible en esta vista)
                    r['ingresos'],
                    r['costo']
                ))
        except Exception as e:
            self._lbl_rent_estado.setText(f"Error al cargar rentabilidad: {e}")
            return

        if not rows:
            self._lbl_rent_estado.setText("Sin datos de rentabilidad para el período seleccionado.")
            return

        total_margen = sum((float(r[3] or 0) - float(r[4] or 0)) for r in rows)

        for i, r in enumerate(rows):
            ingresos    = float(r[3] or 0)
            costo       = float(r[4] or 0)
            margen_monto = ingresos - costo
            margen_pct   = (margen_monto / ingresos * 100) if ingresos > 0 else 0
            contrib_pct  = (margen_monto / total_margen * 100) if total_margen > 0 else 0

            # ABC classification
            if contrib_pct >= 10:    abc = "A ⭐"
            elif contrib_pct >= 3:   abc = "B"
            else:                     abc = "C"

            self._tbl_rent.insertRow(i)
            self._tbl_rent.setItem(i, 0, QTableWidgetItem(str(r[0])))
            self._tbl_rent.setItem(i, 1, QTableWidgetItem(str(r[1] or "")))
            self._tbl_rent.setItem(i, 2, QTableWidgetItem(f"{float(r[2] or 0):.1f}"))
            self._tbl_rent.setItem(i, 3, QTableWidgetItem(f"${ingresos:,.2f}"))
            self._tbl_rent.setItem(i, 4, QTableWidgetItem(f"${costo:,.2f}"))

            item_margen = QTableWidgetItem(f"${margen_monto:,.2f}")
            if margen_pct < 0:
                item_margen.setForeground(__import__('PyQt5.QtGui', fromlist=['QColor']).QColor('#e74c3c'))
            self._tbl_rent.setItem(i, 5, item_margen)

            item_pct = QTableWidgetItem(f"{margen_pct:.1f}%")
            color = '#27ae60' if margen_pct >= 20 else '#e67e22' if margen_pct >= 10 else '#e74c3c'
            item_pct.setForeground(__import__('PyQt5.QtGui', fromlist=['QColor']).QColor(color))
            self._tbl_rent.setItem(i, 6, item_pct)
            self._tbl_rent.setItem(i, 7, QTableWidgetItem(abc))
        self._lbl_rent_estado.setText(f"{len(rows)} productos analizados.")

    def _get_producto_info(self, producto_id: int) -> dict:
        """Obtiene información básica de un producto (nombre, categoría) via AnalyticsEngine."""
        analytics = getattr(self.container, 'analytics_engine', None)
        if analytics and hasattr(analytics, 'get_product_info'):
            try:
                return analytics.get_product_info(producto_id)
            except Exception:
                pass
        # Fallback mínimo sin SQL directo: retornar datos básicos
        return {'nombre': f'Prod {producto_id}', 'categoria': ''}

    def _exportar_rentabilidad_csv(self):
        """Exporta la tabla de rentabilidad a CSV."""
        try:
            from PyQt5.QtWidgets import QFileDialog
            import csv
            path, _ = QFileDialog.getSaveFileName(
                self, "Guardar CSV", "rentabilidad.csv", "CSV (*.csv)")
            if not path: return
            with open(path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                headers = [self._tbl_rent.horizontalHeaderItem(c).text()
                           for c in range(self._tbl_rent.columnCount())]
                writer.writerow(headers)
                for row in range(self._tbl_rent.rowCount()):
                    writer.writerow([
                        self._tbl_rent.item(row, col).text() if self._tbl_rent.item(row, col) else ""
                        for col in range(self._tbl_rent.columnCount())
                    ])
            Toast.success(self, "Exportado", f"CSV guardado:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _crear_tarjeta_kpi(self, titulo, valor_inicial):
        """Crea una tarjeta visual estilizada para los indicadores."""
        # CORRECCIÓN: with_layout=False para evitar conflicto de layouts
        tarjeta = create_card(self, with_layout=False)
        layout = QVBoxLayout(tarjeta)

        lbl_titulo = create_caption(self, titulo)
        lbl_titulo.setAlignment(Qt.AlignCenter)

        lbl_valor = create_subheading(self, valor_inicial)
        lbl_valor.setObjectName("textPrimary")  # Color azul primario
        lbl_valor.setAlignment(Qt.AlignCenter)

        layout.addWidget(lbl_titulo)
        layout.addWidget(lbl_valor)
        tarjeta._lbl_valor = lbl_valor   # keep reference on the frame
        return tarjeta
    def _crear_tabla_ranking(self, titulo, headers):
        """Crea un panel con tabla reusable y filtros de búsqueda."""
        grupo = QGroupBox(titulo)
        grupo.setObjectName("styledGroup")
        layout = QVBoxLayout(grupo)

        tabla = DataTableWithFilters(headers=headers, parent=grupo)
        tabla.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        tabla.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        layout.addWidget(tabla)
        grupo._tabla = tabla
        return grupo

    def _on_global_filters_changed(self, payload: dict) -> None:
        search = (payload or {}).get("search", "").strip().lower()
        wrappers = [self._grp_top._tabla, self._grp_lentos._tabla, self._grp_vips._tabla]
        for wrapper in wrappers:
            wrapper.filter_bar.search.blockSignals(True)
            wrapper.filter_bar.search.setText(search)
            wrapper.filter_bar.search.blockSignals(False)
            wrapper._apply_filter({"search": search})
        self._lbl_rankings_estado.setText(
            "Sin filtros globales activos." if not search else f"Filtro activo: '{search}'"
        )

    def _wire_business_events(self) -> None:
        """Refresco reactivo ante eventos de negocio del ERP."""
        try:
            from core.events.event_bus import get_bus
            from PyQt5.QtCore import QTimer as _QT
            bus = get_bus()
            for evt in ("venta_confirmada", "stock_actualizado", "pago_registrado"):
                bus.subscribe(evt, lambda _p, _self=self: _QT.singleShot(0, _self.cargar_dashboard),
                              label=f"bi_v2.refresh.{evt}")
        except Exception:
            pass

    def _exportar(self, formato: str) -> None:
        """Exporta el dashboard actual a PDF o Excel via ExportService."""
        # [spj-dedup removed local QMessageBox import]
        import os, datetime

        rango = self.cmb_rango.currentText()
        ts    = datetime.datetime.now().strftime("%Y%m%d_%H%M")
        ext   = "xlsx" if formato == "excel" else "pdf"
        nombre_default = f"dashboard_bi_{ts}.{ext}"

        ruta, _ = QFileDialog.getSaveFileName(
            self, f"Guardar {ext.upper()}", nombre_default,
            f"{'Excel (*.xlsx)' if formato == 'excel' else 'PDF (*.pdf)'}"
        )
        if not ruta:
            return
        if not confirm_action(
            self,
            "Confirmar exportación",
            f"¿Deseas exportar el dashboard en formato {ext.upper()}?",
            confirm_text="Sí, exportar",
            cancel_text="Cancelar",
        ):
            return

        try:
            from core.services.export_service import ExportService
            svc  = ExportService(self.container.db)
            # BI unificado: fuente única analytics_engine
            analytics = getattr(self.container, 'analytics_engine', None)
            if not analytics:
                raise RuntimeError("AnalyticsEngine no disponible en el contenedor")
            data = analytics.get_dashboard_data(
                self.sucursal_id,
                rango.lower().split(" ")[-1]
            )
            if formato == "excel":
                sheets = {
                    "KPIs": [data["kpis"]],
                    "Top Productos":    data.get("top_productos", []),
                    "Productos Lentos": data.get("productos_lentos", []),
                    "Clientes VIP":     data.get("clientes_recurrentes", []),
                }
                result = svc.export_ventas(fmt="xlsx")
                # Write our structured BI export
                import openpyxl
                wb = openpyxl.Workbook()
                for sheet_name, rows in sheets.items():
                    ws = wb.create_sheet(title=sheet_name[:31])
                    if rows:
                        ws.append(list(rows[0].keys()))
                        for row in rows:
                            ws.append([str(v) for v in row.values()])
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
                wb.save(ruta)
            else:
                # PDF via ReportEngine if available
                try:
                    from core.services.enterprise.report_engine import ReportEngine
                    engine = ReportEngine(self.container.db)
                    engine.export_pdf("dashboard", data, filepath=ruta)
                except Exception:
                    # Simple PDF fallback
                    svc.export(
                        "SELECT folio, total, fecha FROM ventas "
                        "WHERE DATE(fecha)=DATE('now') ORDER BY fecha DESC LIMIT 500",
                        (), fmt="pdf", filepath=ruta
                    )
            Toast.success(self, "Exportado", f"Archivo guardado en:\n{ruta}")
            os.startfile(ruta) if os.name == "nt" else None
        except Exception as e:
            QMessageBox.critical(self, "Error al exportar", str(e))

    def cargar_dashboard(self):
        """Pide los datos al AnalyticsEngine (BI unificado) y actualiza la UI."""
        rango_str = self.cmb_rango.currentText().lower().split(' ')[-1] # 'hoy', 'semana', 'mes'
        self.loading_dashboard.setVisible(True)
        
        try:
            # BI unificado: fuente única analytics_engine
            analytics = getattr(self.container, 'analytics_engine', None)
            if not analytics:
                raise RuntimeError("AnalyticsEngine no disponible en el contenedor")
            data = analytics.get_dashboard_data(self.sucursal_id, rango_str)
            self._last_data = data
            
            # 1. Actualizar KPIs
            self.lbl_kpi_ingresos._lbl_valor.setText(f"${data['kpis']['ingresos']:,.2f}")
            self.lbl_kpi_ticket._lbl_valor.setText(f"${data['kpis']['ticket_promedio']:,.2f}")
            self.lbl_kpi_ventas._lbl_valor.setText(str(data['kpis']['tickets']))
            self.lbl_kpi_clientes._lbl_valor.setText(str(data['kpis']['clientes_unicos']))

            # Comparativa vs período anterior
            comp = data.get('comparativa', {})
            if comp and hasattr(self, 'lbl_comparativa'):
                prev_ing = comp.get('ingresos', 0)
                curr_ing = data['kpis'].get('ingresos', 0)
                if prev_ing > 0:
                    diff_pct = (curr_ing - prev_ing) / prev_ing * 100
                    arrow = "▲" if diff_pct >= 0 else "▼"
                    color = "#27ae60" if diff_pct >= 0 else "#e74c3c"
                    self.lbl_comparativa.setText(
                        f"<span style='color:{color}'>{arrow} {abs(diff_pct):.1f}% vs período anterior</span>")
                    self.lbl_comparativa.show()
            
            # 2. Llenar Tabla Top Productos
            self._llenar_tabla(self.tabla_top, data['top_productos'], ['nombre', 'cantidad_vendida', 'ingresos_generados'])
            
            # 3. Llenar Tabla Productos Lentos
            self._llenar_tabla(self.tabla_lentos, data['productos_lentos'], ['nombre', 'cantidad_vendida', 'ingresos_generados'])
            
            # 4. Llenar Tabla Clientes VIP
            self._llenar_tabla(self.tabla_vips, data['clientes_recurrentes'], ['nombre', 'visitas', 'valor_vida'])
            top_n = len(data.get('top_productos', []))
            lentos_n = len(data.get('productos_lentos', []))
            vip_n = len(data.get('clientes_recurrentes', []))
            self._lbl_rankings_estado.setText(
                f"Top: {top_n} | Lentos: {lentos_n} | VIP: {vip_n}"
            )
            alertas = []
            if lentos_n > 0:
                alertas.append(f"{lentos_n} productos lentos")
            if top_n == 0:
                alertas.append("sin productos más vendidos")
            self._lbl_resumen_alertas.setText(
                "Alertas: " + (", ".join(alertas) if alertas else "sin alertas relevantes.")
            )
            self._render_echarts_dashboard(data)

        except PermissionError as e:
            # Si el Feature Flag 'bi_v2' está apagado
            QMessageBox.warning(self, "Módulo Inactivo", str(e))
            self._lbl_rankings_estado.setText(f"Módulo inactivo: {e}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo actualizar el dashboard BI.\nDetalle: {e}")
            self._lbl_rankings_estado.setText(f"Error al cargar rankings: {e}")
            self._lbl_resumen_alertas.setText(f"Alertas: error al cargar datos ({e}).")
        finally:
            self.loading_dashboard.setVisible(False)

    def _render_echarts_dashboard(self, data: dict) -> None:
        """Renderiza un mini dashboard con ECharts dentro de QWebEngineView."""
        if not getattr(self, "_chart_view", None):
            return
        top = data.get('top_productos', [])
        if not top:
            self._chart_empty.show()
            self._chart_view.hide()
            return
        labels = [str(i.get('nombre', 'N/A')) for i in top[:8]]
        values = [float(i.get('ingresos_generados', 0) or 0) for i in top[:8]]
        self._chart_empty.hide()
        self._chart_view.show()
        html = f"""
        <html><head><meta charset='utf-8'>
        <script src='https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js'></script>
        <style>html,body,#c{{height:100%;margin:0;background:#0b1220;color:#e2e8f0;font-family:Inter,Arial;}}</style>
        </head><body><div id='c'></div>
        <script>
        const labels = {labels};
        const values = {values};
        if (window.echarts) {{
          const chart = echarts.init(document.getElementById('c'));
          const option = {{
            title: {{text: 'Top productos por ingresos', left: 'center', textStyle: {{color:'#e2e8f0'}}}},
            tooltip: {{trigger: 'axis'}},
            xAxis: {{type: 'category', data: labels, axisLabel: {{color:'#94a3b8', rotate:20}}}},
            yAxis: {{type: 'value', axisLabel: {{color:'#94a3b8'}}}},
            series: [{{type: 'bar', data: values, itemStyle: {{color:'#3b82f6'}}, barMaxWidth: 38}}]
          }};
          chart.setOption(option);
        }} else {{
          const c = document.getElementById('c');
          const rows = labels.map((name, i) => `<div style="display:flex;justify-content:space-between;border-bottom:1px solid #1f2937;padding:8px 4px;"><span>${{name}}</span><b>$${{Number(values[i]||0).toFixed(2)}}</b></div>`).join('');
          c.innerHTML = `<div style="padding:14px;"><h3 style="margin:0 0 8px 0;">Top productos por ingresos</h3><div style="font-size:12px;color:#94a3b8;margin-bottom:8px;">Modo fallback (ECharts no disponible)</div>${{rows}}</div>`;
        }}
        </script></body></html>
        """
        self._chart_view.setHtml(html)

    def _llenar_tabla(self, tabla: QTableWidget, datos: list, llaves: list):
        if not datos:
            tabla.setRowCount(1)
            tabla.setItem(0, 0, QTableWidgetItem("Sin datos para el período seleccionado."))
            for col in range(1, tabla.columnCount()):
                tabla.setItem(0, col, QTableWidgetItem(""))
            if hasattr(tabla.parent(), "empty_state"):
                tabla.parent().empty_state.show()
            return
        tabla.setRowCount(len(datos))
        if hasattr(tabla.parent(), "empty_state"):
            tabla.parent().empty_state.hide()
        for row, item in enumerate(datos):
            for col, llave in enumerate(llaves):
                valor = item.get(llave, '')
                # Dar formato de moneda si es dinero
                if isinstance(valor, (float, int)) and ('ingresos' in llave or 'valor' in llave):
                    txt = f"${valor:,.2f}"
                else:
                    txt = str(valor)
                    
                celda = QTableWidgetItem(txt)
                if col > 0: celda.setTextAlignment(Qt.AlignCenter)
                tabla.setItem(row, col, celda)
