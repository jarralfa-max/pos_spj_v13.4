
# modulos/productos.py
from modulos.design_tokens import Colors, Spacing
from modulos.ui_components import (
    create_primary_button, create_success_button, create_danger_button,
    create_secondary_button, create_table_button, create_input, create_combo,
    create_subheading, create_caption,
    LoadingIndicator, EmptyStateWidget, PageHeader, Toast,
)
import os
from datetime import datetime
from backend.shared.ids import new_uuid
from modulos.spj_refresh_mixin import RefreshMixin
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QLineEdit,
    QComboBox, QMessageBox, QFormLayout, QDoubleSpinBox, QGroupBox,
    QTableWidget, QTableWidgetItem, QDialog, QDialogButtonBox, QHeaderView,
    QAbstractItemView, QFrame, QSplitter, QGridLayout, QListWidget,
    QListWidgetItem, QCompleter, QDateEdit, QTimeEdit, QTabWidget,
    QCheckBox, QSpinBox, QTextEdit,
    QProgressBar, QFileDialog,
    QProgressDialog, QSizePolicy, QScrollArea
)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QPixmap
import logging
from modulos.dialogs.receta_dialog import DialogoReceta
from core.services.recipes.recipe_service import RecipeService
from modulos.kpi_card import KPICard
from core.services.product_catalog_query_service import get_product_configuration_kpis, get_catalog_filter_ids
from backend.application.commands.product_commands import CreateProductCommand, UpdateProductCommand
from backend.application.queries.product_query_service import ProductQueryService
from backend.application.services.product_catalog_service import ProductCatalogService
from backend.application.services.product_image_service import ProductImageService
from backend.application.use_cases.create_product_use_case import CreateProductUseCase
from backend.application.use_cases.deactivate_product_use_case import DeactivateProductCommand, DeactivateProductUseCase
from backend.application.use_cases.restore_product_use_case import RestoreProductCommand, RestoreProductUseCase
from backend.application.use_cases.update_product_use_case import UpdateProductUseCase
from backend.domain.services.product_type_policy import ProductTypePolicy

logger = logging.getLogger(__name__)


class DialogoProducto(QDialog):
    """
    Formulario Modal Enterprise para Crear/Editar Productos.
    Soporta Productos Simples, Compuestos, Subproductos e Imágenes.
    """
    def __init__(self, container, producto_id=None, parent=None):
        super().__init__(parent)
        self.container = container
        self.db = container.db if hasattr(container, 'db') else container
        self.producto_id = producto_id
        self.ruta_imagen_actual = None
        self.product_query_service = ProductQueryService.from_connection(self.db)
        self.product_type_policy = ProductTypePolicy()
        product_service = ProductCatalogService(self.db)
        self.create_product_use_case = CreateProductUseCase(app_service=product_service)
        self.update_product_use_case = UpdateProductUseCase(app_service=product_service)
        self._product_image_service = ProductImageService()
        
        self.setWindowTitle("Nuevo Producto" if not producto_id else f"Editar Producto #{producto_id}")
        self.setMinimumSize(650, 500)
        self.setModal(True)
        
        self.init_ui()
        if self.producto_id:
            self.cargar_datos_producto()

    def keyPressEvent(self, event):
        """
        Override: Ignore Enter/Return so the dialog does NOT auto-accept.
        Enter is used by scanner HID — we don't want it to trigger Save.
        The user must explicitly click the Save button.
        """
        from PyQt5.QtCore import Qt
        if event.key() in (Qt.Key_Return, Qt.Key_Enter):
            # Absorb the event — do nothing
            event.accept()
            return
        super().keyPressEvent(event)

    def init_ui(self):
        layout_principal = QVBoxLayout(self)
        
        # --- TABS DEL FORMULARIO (FASE 1: producto como configuración) ---
        tabs = QTabWidget()
        tab_general = QWidget()
        tab_config = QWidget()
        tab_precios = QWidget()
        tab_receta = QWidget()
        tab_referencias = QWidget()

        tabs.addTab(tab_general, "General")
        tabs.addTab(tab_config, "Configuración")
        tabs.addTab(tab_precios, "Precios / Costos base")
        tabs.addTab(tab_receta, "Receta")
        tabs.addTab(tab_referencias, "Referencias")

        # ================= CAMPOS BASE =================
        
        self.cmb_tipo = QComboBox()
        self.cmb_tipo.addItems(self.product_query_service.type_labels_es())
        self.cmb_tipo.currentTextChanged.connect(self.al_cambiar_tipo)
        
        self.txt_nombre = QLineEdit()
        self.txt_codigo = QLineEdit()
        self.txt_codigo_barras = QLineEdit()
        
        self.cmb_categoria = QComboBox()
        self.cmb_categoria.setEditable(True)
        self.cargar_categorias()
        
        self.cmb_estado = QComboBox()
        self.cmb_estado.addItems(["Activo", "Inactivo"])

        self.cmb_unidad_venta = QComboBox()
        self.cmb_unidad_venta.addItems(["kg", "pza", "litro", "paquete", "caja"])
        self.cmb_unidad_compra = QComboBox()
        self.cmb_unidad_compra.addItems(["kg", "pza", "litro", "paquete", "caja"])
        self.txt_precio = QDoubleSpinBox()
        self.txt_precio.setRange(0.0, 999999.0)
        self.txt_precio.setPrefix("$ ")
        
        self.txt_costo = QDoubleSpinBox()
        self.txt_costo.setRange(0.0, 999999.0)
        self.txt_costo.setPrefix("$ ")
        
        self.cmb_unidad = QComboBox()  # unidad base / inventario
        self.cmb_unidad.addItems(["kg", "pza", "litro", "paquete", "caja"])
        
        self.txt_stock_minimo = QDoubleSpinBox()
        self.txt_stock_minimo.setRange(0.0, 99999.0)

        # Precio mínimo (protección financiera)
        self.txt_precio_minimo = QDoubleSpinBox()
        self.txt_precio_minimo.setRange(0, 999999)
        self.txt_precio_minimo.setDecimals(2)
        self.txt_precio_minimo.setPrefix("$")
        self.txt_precio_minimo.setToolTip("Precio mínimo de venta. Por debajo de este precio el sistema bloquea el descuento.")

        # ================= TAB GENERAL =================
        layout_general = QHBoxLayout(tab_general)
        form_general = QFormLayout()
        form_general.addRow("Nombre:*", self.txt_nombre)
        form_general.addRow("SKU / Código:", self.txt_codigo)
        form_general.addRow("Código de Barras:", self.txt_codigo_barras)
        form_general.addRow("Categoría:", self.cmb_categoria)
        form_general.addRow("Unidad de venta:", self.cmb_unidad_venta)
        form_general.addRow("Unidad de compra:", self.cmb_unidad_compra)
        form_general.addRow("Unidad base / inventario:", self.cmb_unidad)
        form_general.addRow("Estado:", self.cmb_estado)

        # Columna Derecha: Imagen + descripción
        panel_imagen = QVBoxLayout()
        self.lbl_imagen = QLabel("Sin Imagen")
        self.lbl_imagen.setAlignment(Qt.AlignCenter)
        self.lbl_imagen.setFixedSize(180, 180)
        self.lbl_imagen.setObjectName("imagePlaceholder")
        self.lbl_imagen.setToolTip("Vista previa de la imagen del producto. Haga clic en 'Subir Imagen' para cargar una.")
        
        btn_cargar_img = create_secondary_button(self, "📸 Subir Imagen", "Cargar una imagen desde su computadora")
        btn_cargar_img.clicked.connect(self.cargar_imagen)
        
        btn_quitar_img = create_danger_button(self, "❌ Quitar", "Eliminar la imagen actual del producto")
        btn_quitar_img.clicked.connect(self.quitar_imagen)
        
        panel_imagen.addWidget(self.lbl_imagen)
        panel_imagen.addWidget(btn_cargar_img)
        panel_imagen.addWidget(btn_quitar_img)
        panel_imagen.addStretch()
        
        layout_general.addLayout(form_general, 2)
        layout_general.addLayout(panel_imagen, 1)

        # ================= TAB CONFIGURACIÓN =================
        layout_config = QVBoxLayout(tab_config)
        form_config = QFormLayout()
        form_config.addRow("Tipo de Producto:", self.cmb_tipo)
        self.chk_se_vende = QCheckBox("Se vende")
        self.chk_es_inventariable = QCheckBox("Es inventariable")
        self.chk_permite_receta = QCheckBox("Permite receta")
        self.chk_permite_stock_virtual = QCheckBox("Permite stock virtual")
        self.chk_descuenta_componentes = QCheckBox("Descuenta componentes en venta")
        for chk in (
            self.chk_se_vende, self.chk_es_inventariable, self.chk_permite_receta,
            self.chk_permite_stock_virtual, self.chk_descuenta_componentes
        ):
            chk.setEnabled(False)  # fallback seguro: solo mostrar comportamiento
            layout_config.addWidget(chk)
        layout_config.addLayout(form_config)
        self.lbl_tipo_help = create_caption(self, "")
        layout_config.addWidget(self.lbl_tipo_help)
        self._actualizar_hint_tipo(self.cmb_tipo.currentText())

        # ================= TAB PRECIOS / COSTOS BASE =================
        layout_precios = QVBoxLayout(tab_precios)
        form_precios = QFormLayout()
        form_precios.addRow("Precio venta:", self.txt_precio)
        form_precios.addRow("Precio compra base:", self.txt_costo)
        form_precios.addRow("Precio mínimo:", self.txt_precio_minimo)
        self.lbl_costo_std = create_caption(self, "Costo estándar: —")
        self.lbl_margen = create_caption(self, "Margen esperado: —")
        layout_precios.addLayout(form_precios)
        layout_precios.addWidget(self.lbl_costo_std)
        layout_precios.addWidget(self.lbl_margen)

        # ================= TAB RECETA =================
        lay_receta = QVBoxLayout(tab_receta)
        lay_receta.addWidget(create_caption(
            self,
            "La receta del producto se administra desde Productos > pestaña Receta del módulo."
        ))
        lay_receta.addWidget(create_caption(
            self,
            "SIMPLE y SERVICIO no usan receta. COMPUESTO/PROCESABLE/PRODUCIDO sí permiten receta."
        ))
        lay_receta.addStretch()

        # ================= TAB REFERENCIAS =================
        lay_ref = QVBoxLayout(tab_referencias)
        self.lbl_stock_fisico = create_caption(self, "Stock físico actual: —")
        self.lbl_disponible_venta = create_caption(self, "Disponible venta: —")
        lay_ref.addWidget(self.lbl_stock_fisico)
        lay_ref.addWidget(self.lbl_disponible_venta)
        lay_ref.addWidget(create_caption(self, "Stock y disponibilidad se administran en Inventario."))
        btn_ver_inv = create_secondary_button(self, "📦 Ver en Inventario", "Abrir módulo Inventario para gestión de existencias")
        btn_ver_inv.clicked.connect(lambda: QMessageBox.information(
            self, "Inventario",
            "La gestión de existencias se realiza en el módulo Inventario."
        ))
        lay_ref.addWidget(btn_ver_inv)
        lay_ref.addWidget(QLabel("Stock mínimo de referencia:"))
        lay_ref.addWidget(self.txt_stock_minimo)
        lay_ref.addStretch()
        
        # --- BOTONES DE ACCIÓN ---
        layout_principal.addWidget(tabs)
        
        btn_box = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        # Remove default button so Enter doesn't auto-accept (prevents scanner auto-save)
        save_btn = btn_box.button(QDialogButtonBox.Save)
        if save_btn:
            save_btn.setText("Guardar")
            save_btn.setObjectName("primaryBtn")
            save_btn.setDefault(False)
            save_btn.setAutoDefault(False)
        cancel_btn = btn_box.button(QDialogButtonBox.Cancel)
        if cancel_btn:
            cancel_btn.setText("Cancelar")
            cancel_btn.setObjectName("secondaryBtn")
        btn_box.accepted.connect(self.guardar_producto)
        btn_box.rejected.connect(self.reject)
        layout_principal.addWidget(btn_box)

    def al_cambiar_tipo(self, tipo):
        self._actualizar_hint_tipo(tipo)

    def _actualizar_hint_tipo(self, tipo: str):
        if hasattr(self, "lbl_tipo_help"):
            self.lbl_tipo_help.setText(self.product_query_service.type_help_es(tipo))
        if all(hasattr(self, name) for name in ("chk_se_vende", "chk_es_inventariable", "chk_permite_receta", "chk_permite_stock_virtual", "chk_descuenta_componentes")):
            rules = self.product_query_service.type_rules(tipo)
            self.chk_se_vende.setChecked(bool(rules["is_sellable"]))
            self.chk_es_inventariable.setChecked(bool(rules["is_inventory_tracked"]))
            self.chk_permite_receta.setChecked(bool(rules["allows_recipe"]))
            self.chk_permite_stock_virtual.setChecked(bool(rules["allows_virtual_stock"]))
            self.chk_descuenta_componentes.setChecked(bool(rules["deducts_components_on_sale"]))

    def cargar_categorias(self):
        """Carga las categorías únicas existentes mediante QueryService."""
        try:
            for category in self.product_query_service.list_categories():
                self.cmb_categoria.addItem(category)
        except Exception:
            logger.exception("No se pudieron cargar categorías de productos")

    def cargar_imagen(self):
        """Abre el diálogo para seleccionar una imagen."""
        ruta, _ = QFileDialog.getOpenFileName(self, "Seleccionar Imagen", "", "Imágenes (*.png *.jpg *.jpeg *.webp)")
        if ruta:
            try:
                ruta_destino = self._product_image_service.store_image(ruta)
                self.ruta_imagen_actual = ruta_destino
                self.mostrar_imagen_previa(ruta_destino)
            except Exception as e:
                QMessageBox.warning(self, "Error", f"No se pudo copiar la imagen: {e}")

    def mostrar_imagen_previa(self, ruta):
        if ruta and os.path.exists(ruta):
            pixmap = QPixmap(ruta).scaled(180, 180, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.lbl_imagen.setPixmap(pixmap)
        else:
            self.quitar_imagen()

    def quitar_imagen(self):
        self.ruta_imagen_actual = None
        self.lbl_imagen.clear()
        self.lbl_imagen.setText("Sin Imagen")

    def cargar_datos_producto(self):
        """Si estamos editando, carga los datos actuales del producto mediante QueryService."""
        try:
            p = self.product_query_service.get_product(self.producto_id)
            if p:
                self.txt_nombre.setText(p.get('nombre', ''))
                self.txt_codigo.setText(p.get('codigo', ''))
                self.txt_codigo_barras.setText(p.get('codigo_barras', ''))
                self.cmb_categoria.setCurrentText(p.get('categoria', ''))
                self.txt_precio.setValue(float(p.get('precio') or 0.0))
                self.txt_costo.setValue(float(p.get('precio_compra') or p.get('costo') or 0.0))
                self.cmb_unidad.setCurrentText(p.get('unidad', 'pza'))
                self.cmb_unidad_venta.setCurrentText(p.get('unidad_venta', p.get('unidad', 'pza')))
                self.cmb_unidad_compra.setCurrentText(p.get('unidad_compra', p.get('unidad', 'pza')))
                self.cmb_estado.setCurrentText("Activo" if int(p.get('activo', 1) or 1) == 1 else "Inactivo")
                self.txt_stock_minimo.setValue(float(p.get('stock_minimo') or 0.0))
                if hasattr(self, "txt_precio_minimo"):
                    self.txt_precio_minimo.setValue(float(p.get('precio_minimo_venta') or 0.0))

                rules = ProductTypePolicy.rules_for(p.get('tipo_producto'))
                self.cmb_tipo.setCurrentText(rules.label_es)
                self._actualizar_hint_tipo(self.cmb_tipo.currentText())

                existencia = float(p.get('existencia') or 0.0)
                self.lbl_stock_fisico.setText(f"Stock físico actual: {existencia:.3f} {self.cmb_unidad.currentText()}")
                self.lbl_disponible_venta.setText(f"Disponible venta: {existencia:.3f} {self.cmb_unidad.currentText()}")

                self.ruta_imagen_actual = p.get('imagen_path')
                self.mostrar_imagen_previa(self.ruta_imagen_actual)
        except Exception as e:
            logger.error(f"Error cargando producto {self.producto_id}: {e}")

    def _auto_calcular_precio_minimo(self) -> None:
        """Reservado para SystemSettingsService; no aplica defaults numéricos arbitrarios."""
        return

    def guardar_producto(self):
        nombre = self.txt_nombre.text().strip()
        if not nombre:
            QMessageBox.warning(self, "Validación", "El nombre es obligatorio.")
            return

        duplicate = self.product_query_service.find_duplicate_name(
            nombre, exclude_product_id=self.producto_id if self.producto_id else None
        )
        allow_duplicate_name = False
        if duplicate:
            resp = QMessageBox.question(
                self, "⚠️ Producto similar existe",
                f"Ya existe un producto activo con el nombre '{nombre}'\n"
                f"(Código: {duplicate.get('codigo')}, ID: {duplicate.get('id')})\n\n"
                "¿Deseas guardarlo de todas formas?",
                QMessageBox.Yes | QMessageBox.No)
            if resp != QMessageBox.Yes:
                return
            allow_duplicate_name = True

        command_kwargs = dict(
            operation_id=f"product-{new_uuid()}",
            branch_id=str(getattr(self.container, 'sucursal_id', '') or ''),
            user_name=getattr(self, "usuario_actual", "Sistema") or "Sistema",
            name=nombre,
            sku=self.txt_codigo.text().strip() or None,
            barcode=self.txt_codigo_barras.text().strip(),
            category=self.cmb_categoria.currentText().strip(),
            sale_price=self.txt_precio.value(),
            purchase_price=self.txt_costo.value(),
            minimum_sale_price=getattr(self, "txt_precio_minimo", type("x", (), {"value": lambda s: 0.0})()).value(),
            unit=self.cmb_unidad.currentText(),
            sale_unit=self.cmb_unidad_venta.currentText(),
            purchase_unit=self.cmb_unidad_compra.currentText(),
            minimum_stock=self.txt_stock_minimo.value(),
            product_type=ProductTypePolicy.normalize(self.cmb_tipo.currentText()),
            image_path=self.ruta_imagen_actual,
            active=(self.cmb_estado.currentText() == "Activo"),
            allow_duplicate_name=allow_duplicate_name,
        )
        command = (
            UpdateProductCommand(product_id=self.producto_id, **command_kwargs)
            if self.producto_id
            else CreateProductCommand(**command_kwargs)
        )
        result = (
            self.update_product_use_case.execute(command)
            if self.producto_id
            else self.create_product_use_case.execute(command)
        )
        if not result.success:
            if result.message == "PRODUCT_SKU_DUPLICATE":
                QMessageBox.warning(self, "Código duplicado", "El código capturado ya está en uso por otro producto.")
                return
            QMessageBox.critical(self, "No se pudo guardar", "No fue posible guardar el producto. Revise la información e intente nuevamente.")
            return

        if result.data.get("recipe_pending"):
            QMessageBox.information(
                self, "Receta pendiente",
                "El producto fue guardado con receta permitida.\n\n"
                "⚠️  Aún no tiene una receta activa.\n"
                "Administra la receta desde Productos > Receta antes de procesarlo en ventas o producción.")

        if hasattr(self.container, 'audit_service'):
            accion = "ACTUALIZAR_PRODUCTO" if self.producto_id else "CREAR_PRODUCTO"
            self.container.audit_service.log_change(
                usuario="Sistema", accion=accion, modulo="PRODUCTOS",
                entidad="productos", entidad_id=str(result.entity_id)
            )

        self.producto_id = int(result.entity_id) if result.entity_id and str(result.entity_id).isdigit() else self.producto_id
        self.accept()

# ==============================================================================
# MODULO PRINCIPAL (Centro de Productos y Producción)
# ==============================================================================

class ModuloProductos(QWidget, RefreshMixin):
    """
    Centro de Gestión de Productos, Recetas y Procesamiento de Carne (Despiece).
    Integra todo el ciclo de vida del producto en un solo lugar.
    """
    def __init__(self, container, parent=None):
        super().__init__(parent)
        try:
            self._init_refresh(container, ["PRODUCTO_ACTUALIZADO", "PRODUCTO_CREADO", "COMPRA_REGISTRADA"])
        except Exception:
            logger.exception("No se pudo inicializar refresh de productos")
        self.container = container # 🧠 Recibimos el Cerebro
        # Extraemos la db para mantener compatibilidad si algo lo requiere
        self.conexion = container.db if hasattr(container, 'db') else container
        self.product_query_service = ProductQueryService.from_connection(self.conexion)
        # Sucursal desde el contexto de sesión; sin default arbitrario (regla 23).
        self.sucursal_id = getattr(container, "sucursal_id", "") or ""
        self.usuario_actual = ""
        self._product_catalog_service = ProductCatalogService(self.conexion)
        self._deactivate_product_uc = DeactivateProductUseCase(self._product_catalog_service)
        self._restore_product_uc = RestoreProductUseCase(self._product_catalog_service)

        # ── Scanner de código de barras ───────────────────────────────────────
        # Captura input de lectores HID (teclado-emulado).
        # Si el código existe → selecciona y muestra el producto.
        # Si NO existe       → abre DialogoProducto con el código pre-cargado.
        self._scanner_buffer: str = ""
        self._scanner_timer = QTimer(self)
        self._scanner_timer.setSingleShot(True)
        self._scanner_timer.setInterval(80)
        self._scanner_timer.timeout.connect(self._procesar_scanner_producto)

        self.init_ui()

    def set_sucursal(self, sucursal_id: int, nombre_sucursal: str):
        self.sucursal_id = sucursal_id
        self.cargar_catalogo()

    def set_usuario_actual(self, usuario: str, rol: str):
        self.usuario_actual = usuario

    def _crear_stats_productos(self) -> 'QFrame':
        from PyQt5.QtWidgets import QFrame, QHBoxLayout, QPushButton
        bar = QFrame()
        bar.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        bar.setMinimumHeight(116)
        lay = QHBoxLayout(bar)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(Spacing.SM)
        self._kpi_filter_mode = "all"
        self._kpi_cards = {}
        defs = [
            ("activos", "Productos activos", "📦", "success"),
            ("sin_tipo", "Sin tipo_producto", "🏷️", "warning"),
            ("receta_pendiente", "Receta pendiente", "🧪", "info"),
            ("sin_costo", "Sin costo base", "💲", "danger"),
            ("inactivos", "Inactivos", "⏸", "primary"),
        ]
        for key, title, icon, variant in defs:
            btn = QPushButton()
            btn.setFlat(True)
            btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            btn.setMinimumHeight(108)
            btn.clicked.connect(lambda _, k=key: self._on_kpi_click(k))
            card = KPICard(title, "—", icon, variant, parent=btn)
            self._kpi_cards[key] = card
            h = QHBoxLayout(btn)
            h.setContentsMargins(0, 0, 0, 0)
            h.addWidget(card)
            lay.addWidget(btn)
        return bar

    def _refresh_kpi_productos(self):
        try:
            db = self.container.db if hasattr(self.container, 'db') else self.conexion
            data = get_product_configuration_kpis(db)
            for key, card in self._kpi_cards.items():
                card.set_valor(str(data.get(key, 0)))
        except Exception:
            pass

    def _on_kpi_click(self, key: str):
        self._kpi_filter_mode = key
        if key == "inactivos":
            self.cmb_filtro_estado.setCurrentIndex(1)
            return
        if key == "activos":
            self.cmb_filtro_estado.setCurrentIndex(0)
            return
        self.cmb_filtro_estado.setCurrentIndex(2)
        self.tabs.setCurrentWidget(self.tab_catalogo)
        self.cargar_catalogo()

    def init_ui(self):
        layout_principal = QVBoxLayout(self)
        layout_principal.setContentsMargins(0, 0, 0, 0)
        layout_principal.setSpacing(0)

        self._page_header = PageHeader(
            self,
            title="🥩 Centro de Productos",
            subtitle="Catálogo, procesamiento cárnico y activación por sucursal",
        )
        layout_principal.addWidget(self._page_header)

        # ── Stats bar ─────────────────────────────────────────────────────────
        self._stats_productos = self._crear_stats_productos()
        layout_principal.addWidget(self._stats_productos)
        
        # --- PESTAÑAS DEL MÓDULO ---
        self.tabs = QTabWidget()
        self.tabs.setObjectName("tabWidget")
        
        self.tab_catalogo = QWidget()
        self.tab_receta = QWidget()
        self.tab_sucursales = QWidget()

        self.tabs.addTab(self.tab_catalogo,   "📦 Catálogo de Productos")
        self.tabs.addTab(self.tab_receta,     "🧪 Receta")
        self.tabs.addTab(self.tab_sucursales, "🏪 Activación por Sucursal")

        self.setup_tab_catalogo()
        self.setup_tab_receta_producto()
        self.setup_tab_sucursales()
        
        layout_principal.addWidget(self.tabs)
        self.tabs.currentChanged.connect(self.al_cambiar_pestana)

    def al_cambiar_pestana(self, index):
        if index == 0: self.cargar_catalogo()
        elif index == 1: self._refresh_tab_receta_producto()

    # =========================================================
    # PESTAÑA 1: CATÁLOGO DE PRODUCTOS (CRUD ENTERPRISE)
    # =========================================================
    def setup_tab_catalogo(self):
        layout = QVBoxLayout(self.tab_catalogo)
        
        # ── Barra de búsqueda + filtros ───────────────────────────────────
        filtros_layout = QHBoxLayout()
        self.txt_buscar_prod = create_input(self, "🔍 Buscar por nombre, código o barras...", "Ingrese términos de búsqueda para filtrar productos")
        self.txt_buscar_prod.returnPressed.connect(self.cargar_catalogo)
        
        btn_buscar = create_primary_button(self, "🔍 Buscar", "Ejecutar búsqueda de productos")
        btn_buscar.clicked.connect(self.cargar_catalogo)

        # v13.30: Filtro de categoría
        self.cmb_filtro_cat = QComboBox()
        self.cmb_filtro_cat.addItem("📁 Todas")
        self.cmb_filtro_cat.setMinimumWidth(140)
        self.cmb_filtro_cat.currentIndexChanged.connect(self.cargar_catalogo)
        try:
            for category in self.product_query_service.list_categories():
                self.cmb_filtro_cat.addItem(category)
        except Exception:
            pass

        # v13.30: Filtro de estado
        self.cmb_filtro_estado = QComboBox()
        self.cmb_filtro_estado.addItem("✅ Activos", "active")
        self.cmb_filtro_estado.addItem("❌ Eliminados", "deleted")
        self.cmb_filtro_estado.addItem("📋 Todos", "all")
        self.cmb_filtro_estado.setMinimumWidth(130)
        self.cmb_filtro_estado.currentIndexChanged.connect(self.cargar_catalogo)
        
        btn_nuevo = create_success_button(self, "➕ Nuevo Producto", "Crear un nuevo producto en el catálogo")
        btn_nuevo.clicked.connect(self.abrir_nuevo_producto)
        
        btn_historial_precio = create_secondary_button(self, "📈 Historial Precios", "Ver el historial de cambios de precio del producto seleccionado")
        btn_historial_precio.clicked.connect(self._ver_historial_precio)

        btn_importar = create_secondary_button(self, "📥 Importar Excel", 
            "Importar productos desde Excel (.xlsx)\nColumnas requeridas: nombre, precio\nOpcionales: codigo, codigo_barras, categoria, precio_compra, unidad, stock_minimo")
        btn_importar.clicked.connect(self._importar_excel)

        filtros_layout.addWidget(self.txt_buscar_prod, 2)
        filtros_layout.addWidget(self.cmb_filtro_cat)
        filtros_layout.addWidget(self.cmb_filtro_estado)
        filtros_layout.addWidget(btn_buscar)
        filtros_layout.addWidget(btn_nuevo)
        filtros_layout.addWidget(btn_historial_precio)
        filtros_layout.addWidget(btn_importar)
        layout.addLayout(filtros_layout)
        self._loading_catalogo = LoadingIndicator("Cargando catálogo de productos…", self)
        self._loading_catalogo.hide()
        layout.addWidget(self._loading_catalogo)

        # v13.30: Contador de resultados
        self.lbl_conteo = create_caption(self, "")
        layout.addWidget(self.lbl_conteo)
        
        # Tabla de Catálogo
        self.tabla_productos = QTableWidget()
        self.tabla_productos.setColumnCount(9)
        self.tabla_productos.setHorizontalHeaderLabels(
            ["ID", "Código", "Cód.Barras", "Nombre", "Categoría", "Precio", "Stock", "Estado", "Acciones"])
        self.tabla_productos.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.tabla_productos.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabla_productos.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabla_productos.setAlternatingRowColors(True)
        self.tabla_productos.verticalHeader().setVisible(False)
        layout.addWidget(self.tabla_productos)
        self._empty_catalogo = EmptyStateWidget(
            "Sin productos",
            "No se encontraron productos para la búsqueda/filtros actuales.",
            "📭",
            self,
        )
        self._empty_catalogo.hide()
        layout.addWidget(self._empty_catalogo)

    def setup_tab_receta_producto(self):
        lay = QVBoxLayout(self.tab_receta)
        self._loading_receta = LoadingIndicator("Cargando información de receta…", self)
        self._loading_receta.hide()
        self._lbl_receta_estado = create_subheading(self, "Seleccione un producto en Catálogo.")
        self._lbl_receta_hint = create_caption(self, "La receta permitida depende de tipo_producto.")
        self._btn_receta_abrir = create_primary_button(self, "🛠 Gestionar receta", "Crear o editar receta del producto seleccionado")
        self._btn_receta_abrir.clicked.connect(self._abrir_receta_producto)
        self._btn_receta_desactivar = create_secondary_button(self, "⏸ Desactivar receta", "Desactivar receta activa del producto")
        self._btn_receta_desactivar.clicked.connect(self._desactivar_receta_producto)
        self._btn_receta_simular = create_secondary_button(self, "🧮 Simular", "Vista previa informativa según tipo de producto")
        self._btn_receta_simular.clicked.connect(self._simular_receta_producto)
        self._btn_receta_inv = create_secondary_button(self, "📦 Ver en Inventario", "Consultar disponibilidad/stock en Inventario")
        self._btn_receta_inv.clicked.connect(
            lambda: QMessageBox.information(self, "Inventario", "La disponibilidad y stock se consultan en Inventario.")
        )
        self._tbl_receta_componentes = QTableWidget()
        self._tbl_receta_componentes.setColumnCount(5)
        self._tbl_receta_componentes.setHorizontalHeaderLabels(
            ["Componente/Insumo", "Cantidad", "Unidad", "Rendimiento %", "Merma %"]
        )
        self._tbl_receta_componentes.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self._tbl_receta_componentes.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._tbl_receta_componentes.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._lbl_receta_resumen = create_caption(self, "Resumen contextual: —")
        self._lbl_receta_preview = create_caption(self, "")
        self._empty_receta = EmptyStateWidget(
            "Sin contexto de receta",
            "Seleccione un producto en Catálogo para ver o gestionar su receta.",
            "🧪",
            self,
        )
        lay.addWidget(self._lbl_receta_estado)
        lay.addWidget(self._lbl_receta_hint)
        lay.addWidget(self._loading_receta)
        lay.addWidget(self._btn_receta_abrir)
        row_btn = QHBoxLayout()
        row_btn.addWidget(self._btn_receta_simular)
        row_btn.addWidget(self._btn_receta_desactivar)
        row_btn.addWidget(self._btn_receta_inv)
        row_btn.addStretch()
        lay.addLayout(row_btn)
        lay.addWidget(self._tbl_receta_componentes)
        lay.addWidget(self._lbl_receta_resumen)
        lay.addWidget(self._lbl_receta_preview)
        lay.addWidget(self._empty_receta)
        lay.addStretch()
        self._refresh_tab_receta_producto()

    def _producto_seleccionado_catalogo(self):
        if not hasattr(self, "tabla_productos"):
            return None
        row = self.tabla_productos.currentRow()
        if row < 0:
            return None
        try:
            pid = int(self.tabla_productos.item(row, 0).text())
            p = self.product_query_service.get_product(pid)
            if not p:
                return None
            return {"id": p.get("id"), "nombre": p.get("nombre"), "tipo_producto": p.get("tipo_producto")}
        except Exception:
            return None

    def _refresh_tab_receta_producto(self):
        p = self._producto_seleccionado_catalogo()
        self._tbl_receta_componentes.setRowCount(0)
        self._lbl_receta_resumen.setText("Resumen contextual: —")
        self._lbl_receta_preview.setText("")
        if not p:
            self._lbl_receta_estado.setText("Seleccione un producto en Catálogo.")
            self._lbl_receta_hint.setText("SIMPLE: sin receta. COMPUESTO/PROCESABLE/PRODUCIDO: con receta.")
            self._btn_receta_abrir.setEnabled(False)
            self._btn_receta_desactivar.setEnabled(False)
            self._btn_receta_simular.setEnabled(False)
            self._empty_receta.show()
            return
        self._empty_receta.hide()
        tipo = (p.get("tipo_producto") or "simple").lower()
        self._btn_receta_abrir.setEnabled(tipo != "simple" and tipo != "servicio")
        self._btn_receta_simular.setEnabled(tipo in {"compuesto", "procesable", "producido"})
        self._btn_receta_desactivar.setEnabled(tipo in {"compuesto", "procesable", "producido", "subproducto"})
        rules = {
            "simple": "Este producto no usa receta.",
            "compuesto": "Receta de combinación: al vender 1 unidad se descuentan componentes.",
            "procesable": "Receta de despiece: al procesar se generan subproductos.",
            "producido": "Receta de producción: se elabora consumiendo insumos/subproductos.",
            "subproducto": "Este producto es generado por despiece o usado como componente.",
            "insumo": "Este producto se usa como insumo.",
            "servicio": "Los servicios no controlan inventario ni receta.",
        }
        self._lbl_receta_estado.setText(f"Producto: {p['nombre']} ({tipo.upper()})")
        self._lbl_receta_hint.setText(rules.get(tipo, "Revise configuración de receta para este tipo."))
        self._cargar_detalle_receta_producto(p, tipo)

    def _cargar_detalle_receta_producto(self, producto: dict, tipo: str) -> None:
        db = self.container.db if hasattr(self.container, 'db') else self.conexion
        svc = RecipeService(db)
        self._loading_receta.show()
        receta = svc.get_recipe_for_product(producto["id"])
        if not receta:
            if tipo == "simple":
                self._lbl_receta_resumen.setText("Resumen contextual: producto simple (sin receta).")
                self._lbl_receta_preview.setText("Este producto no usa receta. Los productos simples se compran y venden directamente.")
            elif tipo == "servicio":
                self._lbl_receta_resumen.setText("Resumen contextual: servicio (sin receta).")
                self._lbl_receta_preview.setText("Los servicios no controlan inventario ni receta.")
            else:
                self._lbl_receta_resumen.setText("Resumen contextual: receta pendiente.")
                self._lbl_receta_preview.setText("No hay receta activa para este producto.")
            self._loading_receta.hide()
            return
        componentes = svc.get_recipe_components(receta["id"])
        self._tbl_receta_componentes.setRowCount(len(componentes))
        for i, c in enumerate(componentes):
            vals = [
                str(c.get("component_nombre") or ""),
                f"{float(c.get('cantidad') or 0):.3f}",
                str(c.get("unidad") or ""),
                f"{float(c.get('rendimiento_pct') or 0):.3f}",
                f"{float(c.get('merma_pct') or 0):.3f}",
            ]
            for j, v in enumerate(vals):
                self._tbl_receta_componentes.setItem(i, j, QTableWidgetItem(v))
        if tipo == "compuesto":
            self._lbl_receta_resumen.setText(f"Resumen: Componentes {len(componentes)} · Disponible por componentes: — · Componente limitante: —")
            self._lbl_receta_preview.setText("Preview: Al vender 1 unidad se descontará la lista de componentes.")
        elif tipo == "procesable":
            total_r = sum(float(c.get('rendimiento_pct') or 0) for c in componentes)
            total_m = sum(float(c.get('merma_pct') or 0) for c in componentes)
            self._lbl_receta_resumen.setText(
                f"Resumen: Subproductos {len(componentes)} · Total generado {total_r:.1f}% · Merma {total_m:.1f}% · Total receta {(total_r+total_m):.1f}%"
            )
            self._lbl_receta_preview.setText("Preview: Si procesas X kg se generarán subproductos según rendimiento y merma.")
        elif tipo == "producido":
            self._lbl_receta_resumen.setText(f"Resumen: Insumos {len(componentes)} · Costo estimado: — · Rendimiento esperado: —")
            self._lbl_receta_preview.setText("Preview: Si produces X se consumirán insumos según cantidades de la receta.")
        elif tipo == "subproducto":
            usados = self._buscar_usos_subproducto(svc, producto["id"])
            self._lbl_receta_resumen.setText(f"Resumen: Recetas origen: {(1 if receta else 0)} · Usado en: {len(usados)} · Último costo: —")
            self._lbl_receta_preview.setText(f"Usado en {len(usados)} receta(s): {', '.join(usados[:5])}" if usados else "Sin usos detectados en recetas activas.")
        self._loading_receta.hide()

    def _buscar_usos_subproducto(self, svc: RecipeService, producto_id: int):
        usos = []
        for rec in svc.get_all_recipes(include_inactive=False):
            comps = svc.get_recipe_components(rec["id"])
            if any(int(c.get("component_product_id") or 0) == int(producto_id) for c in comps):
                usos.append(str(rec.get("nombre_receta") or f"Receta {rec.get('id')}"))
        return usos

    def _abrir_receta_producto(self):
        p = self._producto_seleccionado_catalogo()
        if not p:
            QMessageBox.information(self, "Receta", "Seleccione un producto en Catálogo.")
            return
        tipo = (p.get("tipo_producto") or "simple").lower()
        if tipo == "simple":
            QMessageBox.information(self, "Receta no permitida", "La UI no permite crear receta para producto SIMPLE.")
            return
        if tipo == "servicio":
            QMessageBox.information(self, "Receta no permitida", "Los servicios no usan receta.")
            return
        db = self.container.db if hasattr(self.container, 'db') else self.conexion
        svc = RecipeService(db)
        prods = svc.get_products_for_ui()
        receta, comps = None, []
        existente = svc.get_recipe_for_product(p["id"])
        if existente:
            receta, comps = svc.get_recipe_data_for_edit(existente["id"])
        dlg = DialogoReceta(svc, prods, getattr(self, "usuario_actual", "Sistema"),
                            receta_data=receta, componentes=comps, parent=self)
        tipo_target = {"compuesto": "COMBINACION", "procesable": "SUBPRODUCTO", "producido": "PRODUCCION", "subproducto": "SUBPRODUCTO"}.get(tipo, "SUBPRODUCTO")
        i_tipo = dlg._combo_tipo_receta.findData(tipo_target)
        if i_tipo >= 0:
            dlg._combo_tipo_receta.setCurrentIndex(i_tipo)
        i_base = dlg._combo_base.findData(p["id"])
        if i_base >= 0:
            dlg._combo_base.setCurrentIndex(i_base)
        dlg._combo_tipo_receta.setEnabled(False)
        dlg._combo_base.setEnabled(False)
        if dlg.exec_() == QDialog.Accepted:
            self._refresh_tab_receta_producto()

    def _desactivar_receta_producto(self):
        p = self._producto_seleccionado_catalogo()
        if not p:
            return
        db = self.container.db if hasattr(self.container, 'db') else self.conexion
        svc = RecipeService(db)
        receta = svc.get_recipe_for_product(p["id"])
        if not receta:
            QMessageBox.information(self, "Receta", "No hay receta activa para desactivar.")
            return
        if QMessageBox.question(self, "Desactivar receta", "¿Desea desactivar la receta activa de este producto?") != QMessageBox.Yes:
            return
        svc.deactivate_recipe(receta["id"], getattr(self, "usuario_actual", "Sistema"))
        self._refresh_tab_receta_producto()

    def _simular_receta_producto(self):
        p = self._producto_seleccionado_catalogo()
        if not p:
            return
        tipo = (p.get("tipo_producto") or "simple").lower()
        if tipo == "compuesto":
            QMessageBox.information(self, "Simulación", "Al vender 1 unidad se descontarán los componentes mostrados.")
        elif tipo == "procesable":
            QMessageBox.information(self, "Simulación", "Si procesas X kg se generarán subproductos conforme a rendimiento/merma.")
        elif tipo == "producido":
            QMessageBox.information(self, "Simulación", "Si produces X se consumirán insumos según la receta.")

    def _on_refresh(self, event_type: str, data: dict) -> None:
        """Auto-refresh catalog on product or purchase events."""
        try: self.cargar_catalogo()
        except Exception as e: logger.debug("refresh error: %s", e)

    def cargar_catalogo(self):
        if hasattr(self, "_loading_catalogo"):
            self._loading_catalogo.show()
        busqueda = self.txt_buscar_prod.text().strip()

        filtro_cat = ""
        if hasattr(self, 'cmb_filtro_cat'):
            cat_text = self.cmb_filtro_cat.currentText()
            if not cat_text.startswith("📁"):
                filtro_cat = cat_text

        filtro_estado = "active"
        if hasattr(self, 'cmb_filtro_estado'):
            filtro_estado = self.cmb_filtro_estado.currentData() or "active"

        try:
            rows = self.product_query_service.list_catalog_rows(
                search=busqueda,
                category=filtro_cat,
                status_filter=filtro_estado,
                limit=1000,
            )
            db = self.container.db if hasattr(self.container, 'db') else self.conexion
            kpi_mode = getattr(self, "_kpi_filter_mode", "all")
            if kpi_mode in {"sin_tipo", "receta_pendiente", "sin_costo"}:
                ids = get_catalog_filter_ids(db, kpi_mode)
                rows = [r for r in rows if int(r['id']) in ids]

            self.tabla_productos.setRowCount(0)
            from PyQt5.QtGui import QColor as _QC
            from PyQt5.QtWidgets import QHBoxLayout as _HL, QWidget as _QW

            for row_idx, row_data in enumerate(rows):
                self.tabla_productos.insertRow(row_idx)
                prod_id = row_data['id']
                activo = int(row_data.get('activo', 1) or 1)
                is_deleted = not activo

                bg_color = _QC(Colors.DANGER.BG_SOFT) if is_deleted else None

                self.tabla_productos.setItem(row_idx, 0, QTableWidgetItem(str(prod_id)))
                self.tabla_productos.setItem(row_idx, 1, QTableWidgetItem(str(row_data.get('codigo') or '')))
                self.tabla_productos.setItem(row_idx, 2, QTableWidgetItem(str(row_data.get('codigo_barras') or '')))
                self.tabla_productos.setItem(row_idx, 3, QTableWidgetItem(str(row_data.get('nombre') or '')))
                self.tabla_productos.setItem(row_idx, 4, QTableWidgetItem(str(row_data.get('categoria') or '')))
                self.tabla_productos.setItem(row_idx, 5, QTableWidgetItem(f"${float(row_data.get('precio') or 0):.2f}"))
                self.tabla_productos.setItem(row_idx, 6, QTableWidgetItem(f"{float(row_data.get('existencia') or 0):.3f}"))

                estado_txt = "✅ Activo" if activo else "❌ Eliminado"
                estado_item = QTableWidgetItem(estado_txt)
                estado_item.setForeground(_QC(Colors.SUCCESS_BASE if activo else Colors.DANGER_HOVER))
                self.tabla_productos.setItem(row_idx, 7, estado_item)

                if bg_color:
                    for ci in range(8):
                        it = self.tabla_productos.item(row_idx, ci)
                        if it:
                            it.setBackground(bg_color)
                            it.setForeground(_QC(Colors.NEUTRAL.SLATE_400))

                _cell = _QW(); _lay = _HL(_cell)
                _lay.setContentsMargins(2, 2, 2, 2); _lay.setSpacing(2)

                btn_editar = create_table_button(self, "✏️", "Editar producto", "outline")
                btn_editar.setFixedSize(28, 26)
                btn_editar.clicked.connect(lambda _, pid=prod_id: self.abrir_editar_producto(pid))
                _lay.addWidget(btn_editar)

                if activo:
                    btn_toggle = create_table_button(self, "🙈", "Ocultar del POS", "warning")
                    btn_toggle.setFixedSize(28, 26)
                    btn_toggle.clicked.connect(lambda _, pid=prod_id, a=activo: self._toggle_activo(pid, a))
                    _lay.addWidget(btn_toggle)

                    btn_del = create_table_button(self, "🗑️", "Eliminar (soft delete)", "danger")
                    btn_del.setFixedSize(28, 26)
                    btn_del.clicked.connect(lambda _, pid=prod_id, nom=row_data.get('nombre', ''): self.eliminar_producto(pid, nom))
                    _lay.addWidget(btn_del)
                else:
                    btn_restaurar = create_table_button(self, "♻️", "Restaurar producto", "success")
                    btn_restaurar.setFixedSize(28, 26)
                    btn_restaurar.clicked.connect(lambda _, pid=prod_id, nom=row_data.get('nombre', ''): self._restaurar_producto(pid, nom))
                    _lay.addWidget(btn_restaurar)

                self.tabla_productos.setCellWidget(row_idx, 8, _cell)

            if hasattr(self, 'lbl_conteo'):
                total = len(rows)
                activos = sum(1 for r in rows if int(r.get('activo', 1) or 1))
                self.lbl_conteo.setText(f"Mostrando {total} productos ({activos} activos, {total - activos} eliminados)")
            if hasattr(self, "_empty_catalogo"):
                self._empty_catalogo.setVisible(len(rows) == 0)
            self._refresh_kpi_productos()

        except Exception as e:
            logger.error(f"Error cargando catálogo: {e}")
            if hasattr(self, "_empty_catalogo"):
                self._empty_catalogo.setVisible(True)
        finally:
            if hasattr(self, "_loading_catalogo"):
                self._loading_catalogo.hide()

    def abrir_nuevo_producto(self):
        # v13.30: Verificar permiso
        try:
            from core.permissions import verificar_permiso
            if not verificar_permiso(self.container, "PRODUCTOS.crear", self):
                return
        except Exception:
            logger.exception("No se pudo verificar permiso PRODUCTOS.crear")
            return
        dlg = DialogoProducto(self.container, parent=self)
        if dlg.exec_() == QDialog.Accepted:
            self.cargar_catalogo()

    def abrir_editar_producto(self, producto_id):
        try:
            from core.permissions import verificar_permiso
            if not verificar_permiso(self.container, "PRODUCTOS.editar", self):
                return
        except Exception:
            logger.exception("No se pudo verificar permiso PRODUCTOS.editar")
            return
        dlg = DialogoProducto(self.container, producto_id=producto_id, parent=self)
        if dlg.exec_() == QDialog.Accepted:
            self.cargar_catalogo()

    def eliminar_producto(self, producto_id, nombre):
        """Aplica un SOFT DELETE. Nunca borramos registros con historial financiero."""
        # v13.30: Verificar permiso
        try:
            from core.permissions import verificar_permiso
            if not verificar_permiso(self.container, "PRODUCTOS.eliminar", self):
                return
        except Exception:
            logger.exception("No se pudo verificar permiso PRODUCTOS.eliminar")
            return
        resp = QMessageBox.question(
            self, "Confirmar Borrado", 
            f"¿Está seguro de eliminar el producto '{nombre}'?\n(Se ocultará del catálogo pero se mantendrá en el historial).",
            QMessageBox.Yes | QMessageBox.No
        )
        if resp == QMessageBox.Yes:
            try:
                self._deactivate_product_uc.execute(
                    DeactivateProductCommand(
                        product_id=int(producto_id),
                        operation_id=f"product-deactivate-{new_uuid()}",
                        user_name=self.usuario_actual or "sistema",
                    )
                )
                Toast.success(self, "Producto eliminado", "El producto se eliminó correctamente.")
                self.cargar_catalogo()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"No se pudo eliminar: {e}")

    def ejecutar_produccion(self):
        if self.cmb_receta_ejecutar.currentIndex() == -1:
            QMessageBox.warning(self, "Aviso", "Seleccione una receta primero.")
            return
            
        receta_id = self.cmb_receta_ejecutar.currentData()
        peso_entrada = self.txt_peso_entrada.value()
        
        resp = QMessageBox.question(
            self, "Confirmar Despiece", 
            f"¿Procesar {peso_entrada} kg de materia prima?\nEsto descontará el producto base y generará los subproductos en el inventario.",
            QMessageBox.Yes | QMessageBox.No
        )
        if resp == QMessageBox.No: return

        try:
            if hasattr(self.container, 'production_service'):
                resultado = self.container.production_service.execute_production(
                    recipe_id=receta_id, input_qty=peso_entrada, branch_id=self.sucursal_id, user_id=self.usuario_actual
                )
                Toast.success(self, "Despiece completado", f"Lote: {resultado.get('folio', 'N/A')}")
                self.txt_peso_entrada.setValue(0)
            else:
                QMessageBox.warning(self, "Aviso", "ProductionService no está conectado aún.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Ocurrió un error:\n{e}")

    def setup_tab_sucursales(self) -> None:
        """
        Pestaña para gestionar activación de productos por sucursal.
        Permite: activar/desactivar, precio local, stock mínimo local.
        """
        from PyQt5.QtWidgets import (
            QVBoxLayout, QHBoxLayout, QLabel, QComboBox, QPushButton,
            QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
            QDoubleSpinBox, QMessageBox, QGroupBox
        )
        lay = QVBoxLayout(self.tab_sucursales)
        lay.setContentsMargins(12, 10, 12, 10); lay.setSpacing(10)

        # ── Info ──────────────────────────────────────────────────────────────
        lbl_info = QLabel(
            "Controla en qué sucursales está disponible cada producto, "
            "su precio local y su stock mínimo local. "
            "Precio/stock en blanco = usar valor global del catálogo."
        )
        lbl_info.setWordWrap(True)
        lbl_info.setObjectName("textSecondary")
        lay.addWidget(lbl_info)

        # ── Filtro por sucursal ───────────────────────────────────────────────
        top = QHBoxLayout()
        top.addWidget(QLabel("Ver sucursal:"))
        self._combo_suc_filter = create_combo(self, ["— Todas —"])
        self._combo_suc_filter.addItem("— Todas —", None)
        self._cargar_sucursales_combo_bp()
        self._combo_suc_filter.currentIndexChanged.connect(self._cargar_tabla_branch_products)
        top.addWidget(self._combo_suc_filter)
        top.addStretch()
        btn_refresh = create_secondary_button(self, "🔄 Actualizar", "Recargar la tabla de productos por sucursal")
        btn_refresh.clicked.connect(self._cargar_tabla_branch_products)
        top.addWidget(btn_refresh)
        lay.addLayout(top)

        # ── Tabla principal ───────────────────────────────────────────────────
        self._tbl_bp = QTableWidget()
        self._tbl_bp.setColumnCount(7)
        self._tbl_bp.setHorizontalHeaderLabels([
            "Sucursal", "Producto", "Activa", "Precio Local", "Stock Mín. Local",
            "Precio Global", "Sucursales inactivas"
        ])
        self._tbl_bp.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._tbl_bp.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._tbl_bp.verticalHeader().setVisible(False)
        self._tbl_bp.setAlternatingRowColors(True)
        hdr = self._tbl_bp.horizontalHeader()
        hdr.setSectionResizeMode(6, QHeaderView.Stretch)
        for i in range(6): hdr.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        lay.addWidget(self._tbl_bp)

        # ── Acciones ──────────────────────────────────────────────────────────
        grp = QGroupBox("Editar selección")
        grp_lay = QHBoxLayout(grp)
        grp_lay.addWidget(QLabel("Activo:"))
        self._combo_bp_activo = QComboBox()
        self._combo_bp_activo.addItems(["✅ Sí", "❌ No"])
        grp_lay.addWidget(self._combo_bp_activo)
        grp_lay.addWidget(QLabel("  Precio local ($):"))
        self._spin_bp_precio = QDoubleSpinBox()
        self._spin_bp_precio.setRange(0, 999999); self._spin_bp_precio.setDecimals(2)
        self._spin_bp_precio.setSpecialValueText("(global)")
        grp_lay.addWidget(self._spin_bp_precio)
        grp_lay.addWidget(QLabel("  Stock mínimo:"))
        self._spin_bp_stock_min = QDoubleSpinBox()
        self._spin_bp_stock_min.setRange(0, 999999); self._spin_bp_stock_min.setDecimals(3)
        self._spin_bp_stock_min.setSpecialValueText("(global)")
        grp_lay.addWidget(self._spin_bp_stock_min)
        grp_lay.addStretch()
        btn_guardar_bp = create_success_button(self, "💾 Guardar Cambios", "Guardar los cambios de activación, precio y stock mínimo del producto seleccionado")
        btn_guardar_bp.clicked.connect(self._guardar_branch_product)
        grp_lay.addWidget(btn_guardar_bp)
        lay.addWidget(grp)

        # Cargar datos iniciales
        self._cargar_tabla_branch_products()

    def _cargar_sucursales_combo_bp(self) -> None:
        try:
            conn = self.conexion
            rows = conn.execute(
                "SELECT id, nombre FROM sucursales WHERE activa=1 ORDER BY id"
            ).fetchall()
            for r in rows:
                self._combo_suc_filter.addItem(f"🏪 {r[1]}", r[0])
        except Exception as e:
            pass

    def _cargar_tabla_branch_products(self) -> None:
        from PyQt5.QtWidgets import QTableWidgetItem
        from PyQt5.QtGui import QColor
        from PyQt5.QtCore import Qt
        try:
            branch_id = self._combo_suc_filter.currentData()
            conn = self.conexion

            if branch_id:
                rows = conn.execute("""
                    SELECT s.nombre as suc_nombre, p.nombre as prod_nombre,
                           bp.activo, bp.precio_local, bp.stock_min_local,
                           p.precio as precio_global,
                           bp.branch_id, bp.product_id
                    FROM branch_products bp
                    JOIN sucursales s ON s.id = bp.branch_id
                    JOIN productos   p ON p.id = bp.product_id
                    WHERE bp.branch_id = ? AND p.activo = 1
                    ORDER BY p.nombre
                """, (branch_id,)).fetchall()
            else:
                rows = conn.execute("""
                    SELECT s.nombre as suc_nombre, p.nombre as prod_nombre,
                           bp.activo, bp.precio_local, bp.stock_min_local,
                           p.precio as precio_global,
                           bp.branch_id, bp.product_id
                    FROM branch_products bp
                    JOIN sucursales s ON s.id = bp.branch_id
                    JOIN productos   p ON p.id = bp.product_id
                    WHERE p.activo = 1
                    ORDER BY s.nombre, p.nombre
                """).fetchall()

            self._tbl_bp.setRowCount(len(rows))
            for ri, r in enumerate(rows):
                # Calcular en qué sucursales está inactivo este producto
                inact = conn.execute("""
                    SELECT GROUP_CONCAT(s2.nombre, ', ')
                    FROM sucursales s2
                    LEFT JOIN branch_products bp2
                        ON bp2.branch_id = s2.id AND bp2.product_id = ?
                    WHERE s2.activa = 1
                      AND (bp2.activo = 0 OR bp2.activo IS NULL)
                """, (r["product_id"],)).fetchone()
                inact_txt = inact[0] if inact and inact[0] else "—"

                vals = [
                    r["suc_nombre"], r["prod_nombre"],
                    "✅ Sí" if r["activo"] else "❌ No",
                    f"${r['precio_local']:.2f}" if r["precio_local"] else "(global)",
                    f"{r['stock_min_local']:.3f}" if r["stock_min_local"] else "(global)",
                    f"${r['precio_global']:.2f}",
                    inact_txt,
                ]
                for ci, v in enumerate(vals):
                    it = QTableWidgetItem(str(v))
                    it.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    if ci == 2 and "No" in str(v):
                        it.setForeground(QColor(Colors.DANGER_HOVER))
                    elif ci == 6 and v != "—":
                        it.setForeground(QColor(Colors.WARNING_BASE))
                    # Store branch_id/product_id for editing
                    if ci == 0:
                        it.setData(Qt.UserRole, (r["branch_id"], r["product_id"]))
                    self._tbl_bp.setItem(ri, ci, it)
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning("_cargar_tabla_branch_products: %s", e)

    def _guardar_branch_product(self) -> None:
        from PyQt5.QtCore import Qt
        # [spj-dedup removed local QMessageBox import]
        row = self._tbl_bp.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Aviso", "Selecciona una fila primero.")
            return
        item_data = self._tbl_bp.item(row, 0)
        if not item_data: return
        branch_id, product_id = item_data.data(Qt.UserRole)
        activo      = 1 if self._combo_bp_activo.currentIndex() == 0 else 0
        precio_local = self._spin_bp_precio.value() if self._spin_bp_precio.value() > 0 else None
        stock_min    = self._spin_bp_stock_min.value() if self._spin_bp_stock_min.value() > 0 else None
        try:
            self.conexion.execute("""
                INSERT INTO branch_products(branch_id, product_id, activo, precio_local, stock_min_local)
                VALUES(?,?,?,?,?)
                ON CONFLICT(branch_id, product_id) DO UPDATE SET
                    activo=excluded.activo,
                    precio_local=excluded.precio_local,
                    stock_min_local=excluded.stock_min_local,
                    updated_at=datetime('now')
            """, (branch_id, product_id, activo, precio_local, stock_min))
            self.conexion.commit()
            Toast.success(self, "Guardado", "Configuración actualizada.")
            self._cargar_tabla_branch_products()
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))


    def _ver_historial_precio(self) -> None:
        """Muestra el historial de cambios de precio del producto seleccionado."""
        from PyQt5.QtWidgets import (
            QDialog, QVBoxLayout, QTableWidget, QTableWidgetItem,
            QHeaderView, QLabel, QPushButton, QMessageBox
        )
        from PyQt5.QtCore import Qt
        from PyQt5.QtGui import QColor

        # Obtener producto seleccionado de la tabla
        tabla = self.tab_catalogo.findChild(QTableWidget)
        if not tabla or tabla.currentRow() < 0:
            Toast.info(self, "Aviso", "Selecciona un producto en la tabla primero.")
            return

        item_id = tabla.item(tabla.currentRow(), 0)
        if not item_id: return
        prod_id  = int(item_id.text()) if item_id.text().isdigit() else None
        if not prod_id: return

        try:
            prod_row = self.conexion.execute(
                "SELECT nombre, precio, precio_compra FROM productos WHERE id=?",
                (prod_id,)
            ).fetchone()
            if not prod_row: return
            nombre_prod = prod_row[0]
            precio_actual = float(prod_row[1] or 0)

            rows = self.conexion.execute("""
                SELECT campo, precio_anterior, precio_nuevo,
                       diferencia_pct, usuario, changed_at
                FROM historial_precios
                WHERE producto_id=?
                ORDER BY changed_at DESC LIMIT 50
            """, (prod_id,)).fetchall()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"No se pudo cargar el historial: {e}")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle(f"📈 Historial de Precios — {nombre_prod}")
        dlg.setMinimumWidth(620)
        lay = QVBoxLayout(dlg)

        lbl = create_subheading(self, f"<b>{nombre_prod}</b>  |  Precio actual: <b>${precio_actual:.2f}</b>")
        lay.addWidget(lbl)

        if not rows:
            lay.addWidget(QLabel(
                "Sin cambios registrados aún.\n"
                "Los cambios se registran automáticamente desde ahora."
            ))
        else:
            tbl = QTableWidget()
            tbl.setColumnCount(6)
            tbl.setHorizontalHeaderLabels([
                "Campo", "Precio anterior", "Precio nuevo",
                "Variación %", "Usuario", "Fecha"
            ])
            tbl.setEditTriggers(0)
            tbl.verticalHeader().setVisible(False)
            tbl.setAlternatingRowColors(True)
            hdr = tbl.horizontalHeader()
            hdr.setSectionResizeMode(5, QHeaderView.Stretch)
            for i in range(5): hdr.setSectionResizeMode(i, QHeaderView.ResizeToContents)
            tbl.setRowCount(len(rows))
            for ri, r in enumerate(rows):
                diff_pct = float(r[3] or 0)
                color = QColor(Colors.DANGER_HOVER) if diff_pct > 10 else                         QColor(Colors.SUCCESS_BASE) if diff_pct < 0 else None
                vals = [
                    str(r[0]), f"${float(r[1]):.2f}", f"${float(r[2]):.2f}",
                    f"{diff_pct:+.1f}%", str(r[4] or "—"), str(r[5] or "")[:16]
                ]
                for ci, v in enumerate(vals):
                    it = QTableWidgetItem(v)
                    it.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    if ci > 0: it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    if ci == 3 and color:
                        it.setForeground(color)
                    tbl.setItem(ri, ci, it)
            lay.addWidget(tbl)

        btn_cerrar = QPushButton("Cerrar")
        btn_cerrar.setObjectName("secondaryBtn")
        btn_cerrar.clicked.connect(dlg.accept)
        lay.addWidget(btn_cerrar)
        dlg.exec_()


    def _importar_excel(self) -> None:
        """Importa productos masivamente desde un archivo Excel (.xlsx)."""
        from PyQt5.QtWidgets import QFileDialog, QProgressDialog
        from PyQt5.QtCore import Qt

        ruta, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar archivo Excel", "",
            "Excel (*.xlsx *.xls *.csv)")
        if not ruta:
            return

        try:
            import openpyxl
        except ImportError:
            QMessageBox.critical(
                self, "Dependencia faltante",
                "Instala openpyxl para importar Excel:\n  pip install openpyxl")
            return

        try:
            wb = openpyxl.load_workbook(ruta, read_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]

            # Mapeo flexible de columnas
            col = {}
            for req, aliases in {
                'nombre':        ['nombre','name','producto','descripcion'],
                'precio':        ['precio','price','precio_venta','pvp'],
                'codigo':        ['codigo','code','sku','clave'],
                'codigo_barras': ['codigo_barras','barcode','ean','upc'],
                'categoria':     ['categoria','category','departamento'],
                'precio_compra': ['precio_compra','costo','cost'],
                'unidad':        ['unidad','unit','um'],
                'stock_minimo':  ['stock_minimo','stock_min','minimo'],
            }.items():
                for alias in aliases:
                    if alias in headers:
                        col[req] = headers.index(alias)
                        break

            if 'nombre' not in col or 'precio' not in col:
                QMessageBox.warning(
                    self, "Formato incorrecto",
                    f"El archivo debe tener columnas 'nombre' y 'precio'.\n"
                    f"Columnas encontradas: {', '.join(headers)}")
                return

            rows = list(ws.iter_rows(min_row=2, values_only=True))
            total = len(rows)
            if total == 0:
                Toast.info(self, "Sin datos", "El archivo no tiene filas de datos.")
                return

            prog = QProgressDialog(f"Importando {total} productos…", "Cancelar", 0, total, self)
            prog.setWindowModality(Qt.WindowModal)

            nuevos = actualizados = errores = 0
            for i, row in enumerate(rows):
                prog.setValue(i)
                if prog.wasCanceled():
                    break
                try:
                    nombre = str(row[col['nombre']] or "").strip()
                    precio = float(row[col['precio']] or 0)
                    if not nombre or precio < 0:
                        errores += 1; continue

                    vals = {
                        'nombre':        nombre,
                        'precio':        precio,
                        'codigo':        str(row[col['codigo']] or "") if 'codigo' in col else "",
                        'codigo_barras': str(row[col['codigo_barras']] or "") if 'codigo_barras' in col else "",
                        'categoria':     str(row[col['categoria']] or "General") if 'categoria' in col else "General",
                        'precio_compra': float(row[col['precio_compra']] or 0) if 'precio_compra' in col else 0.0,
                        'unidad':        str(row[col['unidad']] or "kg") if 'unidad' in col else "kg",
                        'stock_minimo':  float(row[col['stock_minimo']] or 5) if 'stock_minimo' in col else 5.0,
                    }

                    # Check if exists
                    existing = self.container.db.execute(
                        "SELECT id FROM productos WHERE nombre=? OR (codigo!='' AND codigo=?)",
                        (vals['nombre'], vals['codigo'])
                    ).fetchone()

                    if existing:
                        self.container.db.execute("""
                            UPDATE productos SET precio=?, precio_compra=?, categoria=?,
                                unidad=?, stock_minimo=? WHERE id=?
                        """, (vals['precio'], vals['precio_compra'], vals['categoria'],
                              vals['unidad'], vals['stock_minimo'], existing[0]))
                        actualizados += 1
                    else:
                        self.container.db.execute("""
                            INSERT INTO productos
                                (nombre,codigo,codigo_barras,categoria,precio,precio_compra,
                                 unidad,stock_minimo,existencia,activo)
                            VALUES(?,?,?,?,?,?,?,?,0,1)
                        """, (vals['nombre'], vals['codigo'], vals['codigo_barras'],
                              vals['categoria'], vals['precio'], vals['precio_compra'],
                              vals['unidad'], vals['stock_minimo']))
                        nuevos += 1
                except Exception:
                    errores += 1

            self.container.db.commit()
            prog.setValue(total)

            Toast.success(
                self,
                "✅ Importación completa",
                f"Nuevos: {nuevos} · Actualizados: {actualizados} · Errores: {errores}",
            )
            self.cargar_catalogo()

        except Exception as e:
            QMessageBox.critical(self, "Error al importar", str(e))


    # ─────────────────────────────────────────────────────────────────────────
    # SCANNER DE CÓDIGO DE BARRAS
    # Captura input HID (lector emula teclado rápido).
    # Si existe → selecciona fila en tabla. Si no existe → abre diálogo nuevo.
    # ─────────────────────────────────────────────────────────────────────────

    def keyPressEvent(self, event):
        """
        Intercepta el scanner HID SOLO cuando ningún campo de texto tiene el foco.
        Si un QLineEdit/QPlainTextEdit/QSpinBox tiene foco, deja que Qt
        maneje el evento normalmente para no interferir con la edición.
        """
        from PyQt5.QtCore import Qt
        from PyQt5.QtWidgets import QLineEdit, QPlainTextEdit, QDoubleSpinBox, QSpinBox, QTextEdit

        focused = self.focusWidget()
        # If a text-input widget has focus, don't intercept
        if isinstance(focused, (QLineEdit, QPlainTextEdit, QDoubleSpinBox, QSpinBox, QTextEdit)):
            super().keyPressEvent(event)
            return

        text = event.text()
        if event.key() in (Qt.Key_Return, Qt.Key_Enter):
            if self._scanner_buffer:
                self._scanner_timer.stop()
                self._procesar_scanner_producto()
                return
        if text and text.isprintable():
            self._scanner_buffer += text
            self._scanner_timer.start()
            return
        super().keyPressEvent(event)

    def _toggle_activo(self, producto_id: int, activo_actual: int) -> None:
        """Activa o inactiva (oculta del POS) un producto."""
        nuevo = 0 if activo_actual else 1
        label = "activado" if nuevo else "ocultado del POS"
        try:
            self._product_catalog_service.set_product_active(
                product_id=int(producto_id),
                active=bool(nuevo),
                operation_id=f"product-state-{new_uuid()}",
                user_name=self.usuario_actual or "sistema",
            )
            Toast.success(self, "Producto actualizado", f"Producto {label} correctamente.")
            self.cargar_catalogo()
        except Exception as e:
        # [spj-dedup removed local QMessageBox import]
            QMessageBox.critical(self, "Error", str(e))

    def _restaurar_producto(self, producto_id: int, nombre: str) -> None:
        """v13.30: Restaura un producto eliminado (soft delete → activo)."""
        resp = QMessageBox.question(
            self, "♻️ Restaurar Producto",
            f"¿Restaurar el producto '{nombre}'?\n\n"
            "Se volverá a mostrar en el POS y catálogo.",
            QMessageBox.Yes | QMessageBox.No)
        if resp != QMessageBox.Yes:
            return
        try:
            self._restore_product_uc.execute(
                RestoreProductCommand(
                    product_id=int(producto_id),
                    operation_id=f"product-restore-{new_uuid()}",
                    user_name=self.usuario_actual or "sistema",
                )
            )
            Toast.success(self, "✅ Restaurado", f"Producto '{nombre}' restaurado correctamente.")
            self.cargar_catalogo()
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _procesar_scanner_producto(self) -> None:
        """
        Lógica principal del scanner en módulo Productos:
          1. Busca el código en la tabla productos.
          2. Si EXISTE  → selecciona y resalta la fila en la tabla del catálogo.
          3. Si NO existe → abre DialogoProducto con el código pre-cargado.
             El diálogo actual se conserva intacto; solo se pre-llena el campo.
        """
        codigo = self._scanner_buffer.strip()
        self._scanner_buffer = ""
        if not codigo:
            return

        try:
            conn = self.conexion
            row = conn.execute(
                """SELECT id FROM productos
                   WHERE (COALESCE(codigo_barras,'') = ? OR codigo = ?)
                   LIMIT 1""",
                (codigo, codigo)
            ).fetchone()

            if row:
                # Producto encontrado → seleccionar en la tabla
                self._seleccionar_producto_por_id(row[0])
            else:
                # Producto NO encontrado → abrir diálogo nuevo con código pre-cargado
                self._abrir_nuevo_producto_con_codigo(codigo)

        except Exception as e:
            logger.warning("Scanner productos: %s", e)

    def _seleccionar_producto_por_id(self, producto_id: int) -> None:
        """Resalta la fila del producto en la tabla del catálogo."""
        from PyQt5.QtWidgets import QTableWidget
        try:
            # Buscar QTableWidget en el tab_catalogo
            tabla = self.tab_catalogo.findChild(QTableWidget)
            if not tabla:
                return
            # Recargar para asegurar datos frescos
            if hasattr(self, "cargar_catalogo"):
                self.cargar_catalogo()
            for row_idx in range(tabla.rowCount()):
                item = tabla.item(row_idx, 0)
                if item and str(item.text()) == str(producto_id):
                    tabla.selectRow(row_idx)
                    tabla.scrollToItem(item)
                    # Cambiar al tab catálogo
                    if hasattr(self, "tabs"):
                        self.tabs.setCurrentWidget(self.tab_catalogo)
                    break
        except Exception as e:
            logger.warning("_seleccionar_producto_por_id: %s", e)

    def _abrir_nuevo_producto_con_codigo(self, codigo: str) -> None:
        """
        Abre el diálogo DialogoProducto existente con el código de barras
        pre-cargado. NO modifica el diálogo — solo pre-llena el campo.
        """
        from PyQt5.QtWidgets import QDialog
        try:
            dlg = DialogoProducto(self.container, parent=self)
            # Pre-llenar el campo de código de barras con el código escaneado
            if hasattr(dlg, "txt_codigo_barras"):
                dlg.txt_codigo_barras.setText(codigo)
            if dlg.exec_() == QDialog.Accepted:
                if hasattr(self, "cargar_catalogo"):
                    self.cargar_catalogo()
        except Exception as e:
            logger.warning("_abrir_nuevo_producto_con_codigo: %s", e)
    def closeEvent(self, event):
        """Detiene timers activos al cerrar el módulo."""
        try: self._scanner_timer.stop()
        except Exception as e: logger.debug("closeEvent cleanup: %s", e)
        super().closeEvent(event)
