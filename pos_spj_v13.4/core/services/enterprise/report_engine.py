
# core/services/enterprise/report_engine.py
# ── ReportEngine — CEO-Level Enterprise Reports ───────────────────────────────
# All data comes from this engine. No SQL in UI modules.
# Provides: KPI cards, margin real, multi-branch comparison,
#           historical comparison, inventory rotation, loyalty impact.
# Exports: PDF structured, Excel structured.
from __future__ import annotations

import logging
import os
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Dict, List, Optional, Tuple

logger = logging.getLogger("spj.enterprise.reports")


class ReportEngine:

    def __init__(self, db):
        self.db = db

    def _now(self) -> str:
        return datetime.utcnow().isoformat()

    # ── KPI Snapshot ─────────────────────────────────────────────────────────

    def get_kpi_cards(self, branch_id: int,
                       date_from: str, date_to: str) -> Dict:
        """
        Returns executive KPI card data:
        - Total revenue, cost, gross margin, margin %
        - Ticket count, avg ticket
        - Inventory value
        - Active / new clients
        - Points issued
        """
        # Revenue & margin
        sales_row = self.db.fetchone("""
            SELECT
                COALESCE(SUM(v.total),0) AS total_revenue,
                COALESCE(SUM(dv.costo_unitario * dv.cantidad),0) AS total_cost,
                COUNT(DISTINCT v.id) AS ticket_count
            FROM ventas v
            JOIN detalles_venta dv ON dv.venta_id = v.id
            WHERE v.sucursal_id = ?
              AND DATE(v.fecha) BETWEEN DATE(?) AND DATE(?)
              AND v.estado = 'completada'
        """, (branch_id, date_from, date_to))

        total_revenue = float(sales_row["total_revenue"] or 0)
        total_cost    = float(sales_row["total_cost"] or 0)
        ticket_count  = int(sales_row["ticket_count"] or 0)
        gross_margin  = total_revenue - total_cost
        margin_pct    = (gross_margin / total_revenue * 100) if total_revenue else 0
        avg_ticket    = (total_revenue / ticket_count) if ticket_count else 0

        # Inventory value
        inv_row = self.db.fetchone("""
            SELECT COALESCE(SUM(p.existencia * p.precio),0) AS inv_value
            FROM productos p
            WHERE p.is_active = 1
        """)
        inventory_value = float(inv_row["inv_value"] or 0)

        # Clients
        clients_row = self.db.fetchone("""
            SELECT
                COUNT(DISTINCT v.cliente_id) AS active_clients,
                (SELECT COUNT(*) FROM clientes c
                 WHERE c.activo = 1
                   AND DATE(c.fecha_registro) BETWEEN DATE(?) AND DATE(?)
                ) AS new_clients
            FROM ventas v
            WHERE v.sucursal_id = ?
              AND DATE(v.fecha) BETWEEN DATE(?) AND DATE(?)
              AND v.estado = 'completada'
              AND v.cliente_id IS NOT NULL
        """, (date_from, date_to, branch_id, date_from, date_to))

        active_clients = int(clients_row["active_clients"] or 0) if clients_row else 0
        new_clients    = int(clients_row["new_clients"] or 0) if clients_row else 0

        # Loyalty points
        pts_row = self.db.fetchone("""
            SELECT COALESCE(SUM(puntos),0) AS pts_issued
            FROM historico_puntos
            WHERE tipo = 'GANADOS'
              AND DATE(created_at) BETWEEN DATE(?) AND DATE(?)
        """, (date_from, date_to))
        points_issued = int(pts_row["pts_issued"] or 0) if pts_row else 0

        return {
            "branch_id":         branch_id,
            "date_from":         date_from,
            "date_to":           date_to,
            "total_revenue":     round(total_revenue, 2),
            "total_cost":        round(total_cost, 2),
            "gross_margin":      round(gross_margin, 2),
            "gross_margin_pct":  round(margin_pct, 2),
            "ticket_count":      ticket_count,
            "avg_ticket":        round(avg_ticket, 2),
            "inventory_value":   round(inventory_value, 2),
            "active_clients":    active_clients,
            "new_clients":       new_clients,
            "points_issued":     points_issued,
        }

    # ── Multi-branch comparison ───────────────────────────────────────────────

    def get_branch_comparison(self, date_from: str, date_to: str) -> List[Dict]:
        rows = self.db.fetchall("""
            SELECT
                v.sucursal_id,
                COALESCE(SUM(v.total),0) AS revenue,
                COALESCE(SUM(dv.costo_unitario * dv.cantidad),0) AS cost,
                COUNT(DISTINCT v.id) AS tickets,
                COUNT(DISTINCT v.cliente_id) AS clients,
                COALESCE(AVG(v.total),0) AS avg_ticket
            FROM ventas v
            JOIN detalles_venta dv ON dv.venta_id = v.id
            WHERE DATE(v.fecha) BETWEEN DATE(?) AND DATE(?)
              AND v.estado = 'completada'
            GROUP BY v.sucursal_id
            ORDER BY revenue DESC
        """, (date_from, date_to))
        result = []
        for row in rows:
            r = dict(row)
            rev  = float(r["revenue"] or 0)
            cost = float(r["cost"] or 0)
            r["gross_margin"]     = round(rev - cost, 2)
            r["gross_margin_pct"] = round((rev - cost) / rev * 100, 2) if rev else 0
            result.append(r)
        return result

    # ── Historical comparison ─────────────────────────────────────────────────

    def get_historical_comparison(self, branch_id: int,
                                   months: int = 6) -> List[Dict]:
        rows = self.db.fetchall("""
            SELECT
                strftime('%Y-%m', v.fecha) AS year_month,
                COALESCE(SUM(v.total),0) AS revenue,
                COALESCE(SUM(dv.costo_unitario * dv.cantidad),0) AS cost,
                COUNT(DISTINCT v.id) AS tickets
            FROM ventas v
            JOIN detalles_venta dv ON dv.venta_id = v.id
            WHERE v.sucursal_id = ?
              AND v.fecha >= DATE('now', ? || ' months')
              AND v.estado = 'completada'
            GROUP BY year_month
            ORDER BY year_month
        """, (branch_id, f"-{months}"))
        result = []
        for row in rows:
            r = dict(row)
            rev  = float(r["revenue"] or 0)
            cost = float(r["cost"] or 0)
            r["gross_margin"]     = round(rev - cost, 2)
            r["gross_margin_pct"] = round((rev - cost) / rev * 100, 2) if rev else 0
            result.append(r)
        return result

    # ── Inventory rotation ────────────────────────────────────────────────────

    def get_inventory_rotation(self, branch_id: int,
                                 date_from: str, date_to: str) -> List[Dict]:
        rows = self.db.fetchall("""
            SELECT
                p.id,
                p.nombre,
                p.categoria,
                p.existencia AS stock_actual,
                p.unidad,
                COALESCE(SUM(dv.cantidad),0) AS qty_sold,
                COALESCE(SUM(dv.subtotal),0) AS revenue,
                COALESCE(SUM(dv.costo_unitario * dv.cantidad),0) AS cost,
                CASE
                    WHEN p.existencia > 0 AND COALESCE(SUM(dv.cantidad),0) > 0
                    THEN p.existencia / (COALESCE(SUM(dv.cantidad),0) /
                         MAX(1, JULIANDAY(?) - JULIANDAY(?)))
                    ELSE NULL
                END AS days_of_stock
            FROM productos p
            LEFT JOIN detalles_venta dv ON dv.producto_id = p.id
            LEFT JOIN ventas v ON v.id = dv.venta_id
                AND v.sucursal_id = ?
                AND DATE(v.fecha) BETWEEN DATE(?) AND DATE(?)
                AND v.estado = 'completada'
            WHERE p.is_active = 1
            GROUP BY p.id
            ORDER BY qty_sold DESC
        """, (date_to, date_from, branch_id, date_from, date_to))
        return [dict(r) for r in rows]

    # ── Loyalty impact ────────────────────────────────────────────────────────

    def get_loyalty_impact(self, branch_id: int,
                            date_from: str, date_to: str) -> Dict:
        # Revenue from loyalty clients vs non-loyalty
        loyalty_row = self.db.fetchone("""
        SELECT
                COALESCE(SUM(CASE WHEN v.cliente_id IS NOT NULL AND c.puntos > 0
                                  THEN v.total ELSE 0 END),0) AS loyal_revenue,
                COALESCE(SUM(CASE WHEN v.cliente_id IS NULL OR c.puntos = 0
                                  THEN v.total ELSE 0 END),0) AS nonloyal_revenue,
                COUNT(CASE WHEN v.cliente_id IS NOT NULL AND c.puntos > 0 THEN 1 END) AS loyal_tickets,
                COUNT(CASE WHEN v.cliente_id IS NULL OR c.puntos = 0 THEN 1 END) AS nonloyal_tickets
            FROM ventas v
            LEFT JOIN clientes c ON c.id = v.cliente_id
            WHERE v.sucursal_id = ?
              AND DATE(v.fecha) BETWEEN DATE(?) AND DATE(?)
              AND v.estado = 'completada'
        """, (branch_id, date_from, date_to))

        # Community goal progress
        today = date.today().isoformat()
        community_row = self.db.fetchone("""
        SELECT name, target_value, current_value
            FROM loyalty_community_goals
            WHERE is_active = 1
              AND start_date <= ?
              AND end_date >= ?
              AND (branch_id IS NULL OR branch_id = ?)
            LIMIT 1
        """, (today, today, branch_id))

        result = {
            "loyal_revenue":      float(loyalty_row["loyal_revenue"] or 0),
            "nonloyal_revenue":   float(loyalty_row["nonloyal_revenue"] or 0),
            "loyal_tickets":      int(loyalty_row["loyal_tickets"] or 0),
            "nonloyal_tickets":   int(loyalty_row["nonloyal_tickets"] or 0),
            "community_goal":     dict(community_row) if community_row else None,
        }
        total = result["loyal_revenue"] + result["nonloyal_revenue"]
        result["loyal_revenue_pct"] = (
            round(result["loyal_revenue"] / total * 100, 1) if total else 0
        )
        return result

    # ── KPI snapshot persistence ──────────────────────────────────────────────

    def save_daily_snapshot(self, branch_id: int,
                              snapshot_date: Optional[str] = None) -> None:
        if not snapshot_date:
            snapshot_date = date.today().isoformat()
        kpi = self.get_kpi_cards(branch_id, snapshot_date, snapshot_date)
        try:
            self.db.execute("""
                INSERT INTO kpi_snapshots (
                    branch_id, snapshot_date,
                    total_revenue, total_cost, gross_margin, gross_margin_pct,
                    ticket_count, avg_ticket, inventory_value,
                    active_clients, new_clients, points_issued
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(branch_id, snapshot_date)
                DO UPDATE SET
                    total_revenue    = excluded.total_revenue,
                    total_cost       = excluded.total_cost,
                    gross_margin     = excluded.gross_margin,
                    gross_margin_pct = excluded.gross_margin_pct,
                    ticket_count     = excluded.ticket_count,
                    avg_ticket       = excluded.avg_ticket,
                    inventory_value  = excluded.inventory_value,
                    active_clients   = excluded.active_clients,
                    new_clients      = excluded.new_clients,
                    points_issued    = excluded.points_issued,
                    computed_at      = datetime('now')
            """, (
                branch_id, snapshot_date,
                kpi["total_revenue"], kpi["total_cost"],
                kpi["gross_margin"], kpi["gross_margin_pct"],
                kpi["ticket_count"], kpi["avg_ticket"],
                kpi["inventory_value"],
                kpi["active_clients"], kpi["new_clients"],
                kpi["points_issued"],
            ))
        except Exception as exc:
            logger.warning("kpi_snapshot save failed: %s", exc)

    # ── Export ────────────────────────────────────────────────────────────────

    def export_pdf(self, report_type: str, data: Dict,
                    output_path: str, exported_by: str) -> str:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle,
                Paragraph, Spacer
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm

            os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

            doc = SimpleDocTemplate(output_path, pagesize=A4,
                                    topMargin=1.5*cm, bottomMargin=1.5*cm,
                                    leftMargin=2*cm, rightMargin=2*cm)
            styles = getSampleStyleSheet()

            CORP_BLUE    = colors.HexColor("#1A3A5C")
            CORP_ACCENT  = colors.HexColor("#2E86AB")
            CORP_LIGHT   = colors.HexColor("#E8F4F8")
            CORP_WHITE   = colors.white
            CORP_GRAY    = colors.HexColor("#6C757D")

            title_style = ParagraphStyle(
                "CorpTitle",
                parent=styles["Title"],
                textColor=CORP_BLUE,
                fontSize=18,
                spaceAfter=6,
            )
            sub_style = ParagraphStyle(
                "CorpSub",
                parent=styles["Normal"],
                textColor=CORP_GRAY,
                fontSize=10,
            )

            elements = []
            elements.append(Paragraph(
                f"SPJ Enterprise — {report_type}", title_style
            ))
            elements.append(Paragraph(
                f"Generado por {exported_by} el {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC",
                sub_style
            ))
            elements.append(Spacer(1, 0.5*cm))

            # KPI cards table if applicable
            if report_type == "KPI" and "total_revenue" in data:
                kpi_rows = [
                    ["Métrica", "Valor"],
                    ["Ingresos totales",     f"${data['total_revenue']:,.2f}"],
                    ["Costo total",          f"${data['total_cost']:,.2f}"],
                    ["Margen bruto",         f"${data['gross_margin']:,.2f}"],
                    ["Margen %",             f"{data['gross_margin_pct']:.1f}%"],
                    ["Tickets",              str(data['ticket_count'])],
                    ["Ticket promedio",      f"${data['avg_ticket']:,.2f}"],
                    ["Clientes activos",     str(data['active_clients'])],
                    ["Clientes nuevos",      str(data['new_clients'])],
                    ["Puntos emitidos",      str(data['points_issued'])],
                ]
                tbl = Table(kpi_rows, colWidths=[10*cm, 6*cm])
                tbl.setStyle(TableStyle([
                    ("BACKGROUND",   (0, 0), (-1, 0),  CORP_BLUE),
                    ("TEXTCOLOR",    (0, 0), (-1, 0),  CORP_WHITE),
                    ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [CORP_WHITE, CORP_LIGHT]),
                    ("GRID",         (0, 0), (-1, -1),  0.5, CORP_ACCENT),
                    ("FONTSIZE",     (0, 0), (-1, -1),  10),
                    ("LEFTPADDING",  (0, 0), (-1, -1),  8),
                    ("RIGHTPADDING", (0, 0), (-1, -1),  8),
                    ("TOPPADDING",   (0, 0), (-1, -1),  6),
                    ("BOTTOMPADDING",(0, 0), (-1, -1),  6),
                ]))
                elements.append(tbl)

            # Generic table for list data
            elif isinstance(data.get("rows"), list) and data["rows"]:
                headers = list(data["rows"][0].keys())
                table_data = [headers]
                for row in data["rows"]:
                    table_data.append([str(row.get(h, "")) for h in headers])
                tbl = Table(table_data)
                tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0),  CORP_BLUE),
                    ("TEXTCOLOR",  (0, 0), (-1, 0),  CORP_WHITE),
                    ("FONTNAME",   (0, 0), (-1, 0),  "Helvetica-Bold"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [CORP_WHITE, CORP_LIGHT]),
                    ("GRID",       (0, 0), (-1, -1),  0.5, CORP_ACCENT),
                    ("FONTSIZE",   (0, 0), (-1, -1),  9),
                ]))
                elements.append(tbl)

            doc.build(elements)

            self._log_export(report_type, "PDF", output_path, exported_by,
                             len(data.get("rows", [])))
            return output_path

        except ImportError:
            raise RuntimeError("reportlab not installed — cannot export PDF")
        except Exception as exc:
            logger.error("PDF export failed: %s", exc)
            raise

    def export_excel(self, report_type: str, sheets_data: Dict[str, List[Dict]],
                      output_path: str, exported_by: str) -> str:
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )

            os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            CORP_BLUE_HEX   = "1A3A5C"
            CORP_ACCENT_HEX = "2E86AB"
            CORP_LIGHT_HEX  = "E8F4F8"

            for sheet_name, rows in sheets_data.items():
                ws = wb.create_sheet(title=sheet_name[:31])
                if not rows:
                    continue

                headers = list(rows[0].keys())

                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill("solid", fgColor=CORP_BLUE_HEX)
                header_align = Alignment(horizontal="center", vertical="center")

                alt_fill  = PatternFill("solid", fgColor=CORP_LIGHT_HEX)
                border = Border(
                    bottom=Side(style="thin", color=CORP_ACCENT_HEX)
                )

                # Headers
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
                    cell.font  = header_font
                    cell.fill  = header_fill
                    cell.alignment = header_align
                    cell.border = border

                # Data rows
                for row_idx, row in enumerate(rows, 2):
                    fill = alt_fill if row_idx % 2 == 0 else None
                    for col_idx, header in enumerate(headers, 1):
                        val = row.get(header, "")
                        cell = ws.cell(row=row_idx, column=col_idx, value=val)
                        if fill:
                            cell.fill = fill
                        cell.border = border

                # Auto column width
                for col in ws.columns:
                    max_len = max(
                        (len(str(c.value)) for c in col if c.value is not None),
                        default=8
                    )
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

                # Freeze header
                ws.freeze_panes = "A2"

            wb.save(output_path)

            total_rows = sum(len(rows) for rows in sheets_data.values())
            self._log_export(report_type, "EXCEL", output_path, exported_by, total_rows)
            return output_path

        except ImportError:
            raise RuntimeError("openpyxl not installed — cannot export Excel")
        except Exception as exc:
            logger.error("Excel export failed: %s", exc)
            raise

    # ── Internals ─────────────────────────────────────────────────────────────

    def _log_export(self, report_type: str, fmt: str, file_path: str,
                     exported_by: str, row_count: int) -> None:
        try:
            self.db.execute("""
                INSERT INTO report_export_log (
                    report_type, format, exported_by,
                    file_path, row_count, exported_at
                ) VALUES (?,?,?,?,?,?)
            """, (report_type, fmt, exported_by, file_path,
                  row_count, self._now()))
        except Exception as exc:
            logger.warning("export_log write failed: %s", exc)

    # ── Additional query methods (Block 7 supplement) ─────────────────────────

    def get_daily_sales(self, branch_id: int, date_from: str, date_to: str) -> List[Dict]:
        """Daily sales aggregation for the period."""
        rows = self.db.fetchall("""
            SELECT
                DATE(v.fecha)                          AS fecha,
                COALESCE(s.nombre, 'Principal')        AS sucursal_nombre,
                COUNT(DISTINCT v.id)                   AS tickets,
                COALESCE(SUM(v.total),0)               AS ingresos,
                COALESCE(SUM(dv.costo_unitario*dv.cantidad),0) AS costo,
                COALESCE(SUM(v.total)-SUM(dv.costo_unitario*dv.cantidad),0) AS margen
            FROM ventas v
            LEFT JOIN detalles_venta dv ON dv.venta_id = v.id
            LEFT JOIN sucursales s ON s.id = v.sucursal_id
            WHERE v.sucursal_id = ?
              AND DATE(v.fecha) BETWEEN DATE(?) AND DATE(?)
              AND v.estado = 'completada'
            GROUP BY DATE(v.fecha)
            ORDER BY fecha DESC
        """, (branch_id, date_from, date_to))
        result = []
        for r in rows:
            ing = float(r["ingresos"] or 0)
            mar = float(r["margen"] or 0)
            result.append({**dict(r), "margin_pct": round(mar/ing*100 if ing else 0, 2)})
        return result

    def get_top_products(self, branch_id: int, date_from: str, date_to: str, limit: int = 10) -> List[Dict]:
        """Top products by revenue in period."""
        rows = self.db.fetchall("""
            SELECT p.nombre, p.categoria,
                   COALESCE(SUM(dv.cantidad),0)                          AS qty,
                   COALESCE(SUM(dv.subtotal),0)                          AS revenue,
                   COALESCE(SUM(dv.costo_unitario*dv.cantidad),0)        AS cost
            FROM detalles_venta dv
            JOIN ventas v ON v.id = dv.venta_id
            JOIN productos p ON p.id = dv.producto_id
            WHERE v.sucursal_id = ?
              AND DATE(v.fecha) BETWEEN DATE(?) AND DATE(?)
              AND v.estado = 'completada'
            GROUP BY p.id
            ORDER BY revenue DESC
            LIMIT ?
        """, (branch_id, date_from, date_to, limit))
        result = []
        for r in rows:
            rev = float(r["revenue"] or 0); cost = float(r["cost"] or 0)
            result.append({**dict(r), "margin_pct": round((rev-cost)/rev*100 if rev else 0, 2)})
        return result

    def get_product_margins(self, branch_id: int, date_from: str, date_to: str) -> List[Dict]:
        """Per-product margin analysis."""
        rows = self.db.fetchall("""
            SELECT p.nombre,
                   AVG(dv.costo_unitario)                                AS costo_prom,
                   AVG(dv.precio_unitario)                               AS precio_prom,
                   SUM(dv.subtotal - dv.costo_unitario*dv.cantidad)      AS margen_abs,
                   SUM(dv.subtotal)                                      AS revenue,
                   SUM(dv.cantidad)                                      AS qty
            FROM detalles_venta dv
            JOIN ventas v ON v.id = dv.venta_id
            JOIN productos p ON p.id = dv.producto_id
            WHERE v.sucursal_id = ?
              AND DATE(v.fecha) BETWEEN DATE(?) AND DATE(?)
              AND v.estado = 'completada'
            GROUP BY p.id
            ORDER BY margen_abs DESC
        """, (branch_id, date_from, date_to))
        result = []
        for r in rows:
            rev = float(r["revenue"] or 0)
            mar = float(r["margen_abs"] or 0)
            result.append({**dict(r), "margin_pct": round(mar/rev*100 if rev else 0, 2)})
        return result

    def get_margin_anomalies(self, branch_id: int, date_from: str, date_to: str) -> List[Dict]:
        """Margin anomaly alerts from margin_anomalies table."""
        rows = self.db.fetchall("""
            SELECT ma.week_label, ma.branch_id, ma.product_id,
                   ma.negative_margin_pct AS margin_pct, ma.created_at
            FROM margin_anomalies ma
            WHERE ma.branch_id = ?
              AND DATE(ma.created_at) BETWEEN DATE(?) AND DATE(?)
            ORDER BY ma.created_at DESC
            LIMIT 100
        """, (branch_id, date_from, date_to))
        return [dict(r) for r in rows]

    def get_top_loyal_clients(self, branch_id: int, date_from: str, date_to: str, limit: int = 20) -> List[Dict]:
        """Top clients by loyalty and purchase volume."""
        rows = self.db.fetchall("""
            SELECT c.nombre,
                   COALESCE(c.nivel_fidelidad,'Bronce') AS nivel,
                   COALESCE(c.puntos,0) AS puntos,
                   COUNT(DISTINCT v.id) AS visitas,
                   COALESCE(SUM(v.total),0) AS total,
                   COALESCE(SUM(dv.subtotal - dv.costo_unitario*dv.cantidad),0) AS margen
            FROM clientes c
            JOIN ventas v ON v.cliente_id = c.id
            LEFT JOIN detalles_venta dv ON dv.venta_id = v.id
            WHERE v.sucursal_id = ?
              AND DATE(v.fecha) BETWEEN DATE(?) AND DATE(?)
              AND v.estado = 'completada'
            GROUP BY c.id
            ORDER BY total DESC
            LIMIT ?
        """, (branch_id, date_from, date_to, limit))
        return [dict(r) for r in rows]

    def export_pdf(self, branch_id: int, date_from: str, date_to: str,
                   output_path: str, exported_by: str = "Sistema") -> str:
        """Generate structured CEO PDF report."""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
            from reportlab.lib.enums import TA_CENTER, TA_LEFT

            kpis = self.get_kpi_cards(branch_id, date_from, date_to)
            daily = self.get_daily_sales(branch_id, date_from, date_to)
            top_p = self.get_top_products(branch_id, date_from, date_to, limit=20)

            doc = SimpleDocTemplate(output_path, pagesize=landscape(A4),
                                    rightMargin=1.5*cm, leftMargin=1.5*cm,
                                    topMargin=2*cm, bottomMargin=1.5*cm)
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle("title", parent=styles["Heading1"],
                                         textColor=colors.HexColor("#1a252f"),
                                         fontSize=18, spaceAfter=6, alignment=TA_CENTER)
            h2_style = ParagraphStyle("h2", parent=styles["Heading2"],
                                      textColor=colors.HexColor("#2980b9"),
                                      fontSize=12, spaceBefore=12, spaceAfter=4)
            story = []

            # Title
            story.append(Paragraph(f"Reporte CEO — Período {date_from} a {date_to}", title_style))
            story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#2980b9")))
            story.append(Spacer(1, 0.4*cm))

            # KPI section
            story.append(Paragraph("Indicadores Clave de Rendimiento", h2_style))
            kpi_data = [
                ["Métrica","Valor"],
                ["Ingresos Totales", f"$ {kpis.get('total_revenue',0):,.2f}"],
                ["Margen Real", f"{kpis.get('margin_pct',0):.1f}%"],
                ["Tickets Emitidos", f"{int(kpis.get('ticket_count',0)):,}"],
                ["Ticket Promedio", f"$ {kpis.get('avg_ticket',0):,.2f}"],
                ["Valor Inventario", f"$ {kpis.get('inventory_value',0):,.2f}"],
                ["Clientes Activos", str(int(kpis.get('active_clients',0)))],
            ]
            kpi_tbl = Table(kpi_data, colWidths=[6*cm, 6*cm])
            kpi_tbl.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1a252f")),
                ("TEXTCOLOR",(0,0),(-1,0),colors.white),
                ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                ("FONTSIZE",(0,0),(-1,-1),10),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f8f9fa")]),
                ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#dee2e6")),
                ("PADDING",(0,0),(-1,-1),6),
            ]))
            story.append(kpi_tbl); story.append(Spacer(1,0.5*cm))

            # Daily sales
            if daily:
                story.append(Paragraph("Ventas Diarias", h2_style))
                d_headers = ["Fecha","Tickets","Ingresos","Costo","Margen","Margen%"]
                d_data = [d_headers] + [[r["fecha"][:10],str(r["tickets"]),
                    f"$ {float(r.get('ingresos',0)):,.2f}",
                    f"$ {float(r.get('costo',0)):,.2f}",
                    f"$ {float(r.get('margen',0)):,.2f}",
                    f"{float(r.get('margin_pct',0)):.1f}%"] for r in daily[:30]]
                d_tbl = Table(d_data, colWidths=[3.5*cm,2.5*cm,4*cm,4*cm,4*cm,3*cm])
                d_tbl.setStyle(TableStyle([
                    ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#2980b9")),
                    ("TEXTCOLOR",(0,0),(-1,0),colors.white),
                    ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                    ("FONTSIZE",(0,0),(-1,-1),8),
                    ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f8f9fa")]),
                    ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#dee2e6")),
                    ("PADDING",(0,0),(-1,-1),5),
                ]))
                story.append(d_tbl); story.append(Spacer(1,0.5*cm))

            # Top products
            if top_p:
                story.append(Paragraph("Top 20 Productos por Ingresos", h2_style))
                p_headers = ["Producto","Categoría","Unidades","Ingresos","Margen%"]
                p_data = [p_headers] + [[r["nombre"][:30],r["categoria"] or "",
                    f"{float(r.get('qty',0)):,.2f}",f"$ {float(r.get('revenue',0)):,.2f}",
                    f"{float(r.get('margin_pct',0)):.1f}%"] for r in top_p]
                p_tbl = Table(p_data, colWidths=[7*cm,4*cm,3*cm,4*cm,3*cm])
                p_tbl.setStyle(TableStyle([
                    ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#27ae60")),
                    ("TEXTCOLOR",(0,0),(-1,0),colors.white),
                    ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                    ("FONTSIZE",(0,0),(-1,-1),8),
                    ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f0fff4")]),
                    ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#dee2e6")),
                    ("PADDING",(0,0),(-1,-1),5),
                ]))
                story.append(p_tbl)

            doc.build(story)
            self._log_export("CEO", "PDF", output_path, exported_by, len(daily)+len(top_p))
            return output_path
        except ImportError:
            raise RuntimeError("reportlab not installed — cannot export PDF")
        except Exception as exc:
            logger.error("PDF export failed: %s", exc)
            raise

    def export_excel(self, branch_id: int, date_from: str, date_to: str,
                     output_path: str, exported_by: str = "Sistema") -> str:
        """Generate structured CEO Excel workbook."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
            from openpyxl.utils import get_column_letter

            kpis  = self.get_kpi_cards(branch_id, date_from, date_to)
            daily = self.get_daily_sales(branch_id, date_from, date_to)
            top_p = self.get_top_products(branch_id, date_from, date_to, limit=50)
            rot   = self.get_inventory_rotation(branch_id, date_from, date_to)
            top_c = self.get_top_loyal_clients(branch_id, date_from, date_to)

            wb = Workbook()
            hdr_fill  = PatternFill("solid", fgColor="1A252F")
            hdr_font  = Font(color="ECEFF1", bold=True)
            alt_fill  = PatternFill("solid", fgColor="F8F9FA")
            pos_font  = Font(color="27AE60", bold=True)
            neg_font  = Font(color="E74C3C", bold=True)
            def _header_row(ws, headers, fill=hdr_fill, font=hdr_font):
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.fill = fill; cell.font = font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            def _auto_width(ws):
                for col in ws.columns:
                    max_len = max((len(str(c.value)) for c in col if c.value), default=8)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len+4,40)

            # Sheet 1: KPIs
            ws1 = wb.active; ws1.title = "KPIs"
            ws1.append(["Indicador","Valor"])
            _header_row(ws1,["Indicador","Valor"])
            kpi_rows = [("Ingresos Totales",f"$ {kpis.get('total_revenue',0):,.2f}"),
                        ("Margen Real",f"{kpis.get('margin_pct',0):.1f}%"),
                        ("Tickets",int(kpis.get('ticket_count',0))),
                        ("Ticket Promedio",f"$ {kpis.get('avg_ticket',0):,.2f}"),
                        ("Valor Inventario",f"$ {kpis.get('inventory_value',0):,.2f}"),
                        ("Clientes Activos",int(kpis.get('active_clients',0)))]
            for row_idx,(k,v) in enumerate(kpi_rows, 2):
                ws1.cell(row=row_idx, column=1, value=k)
                ws1.cell(row=row_idx, column=2, value=v)
                if row_idx%2==0: ws1.cell(row=row_idx,column=1).fill=alt_fill; ws1.cell(row=row_idx,column=2).fill=alt_fill
            _auto_width(ws1)

            # Sheet 2: Daily Sales
            ws2 = wb.create_sheet("Ventas Diarias")
            hdrs2 = ["Fecha","Sucursal","Tickets","Ingresos","Costo","Margen","Margen%"]
            _header_row(ws2, hdrs2)
            for row_idx,r in enumerate(daily,2):
                ws2.cell(row=row_idx,column=1,value=r.get("fecha","")[:10])
                ws2.cell(row=row_idx,column=2,value=r.get("sucursal_nombre",""))
                ws2.cell(row=row_idx,column=3,value=int(r.get("tickets",0)))
                ws2.cell(row=row_idx,column=4,value=float(r.get("ingresos",0)))
                ws2.cell(row=row_idx,column=5,value=float(r.get("costo",0)))
                ws2.cell(row=row_idx,column=6,value=float(r.get("margen",0)))
                mp_cell = ws2.cell(row=row_idx,column=7,value=float(r.get("margin_pct",0)))
                mp_cell.font = pos_font if float(r.get("margin_pct",0))>=15 else neg_font
                if row_idx%2==0:
                    for c in range(1,8): ws2.cell(row=row_idx,column=c).fill=alt_fill
            _auto_width(ws2)

            # Sheet 3: Top Products
            ws3 = wb.create_sheet("Top Productos")
            _header_row(ws3,["Producto","Categoría","Unidades","Ingresos","Costo","Margen","Margen%"])
            for row_idx,r in enumerate(top_p,2):
                ws3.cell(row=row_idx,column=1,value=r.get("nombre",""))
                ws3.cell(row=row_idx,column=2,value=r.get("categoria",""))
                ws3.cell(row=row_idx,column=3,value=float(r.get("qty",0)))
                ws3.cell(row=row_idx,column=4,value=float(r.get("revenue",0)))
                ws3.cell(row=row_idx,column=5,value=float(r.get("cost",0)))
                ws3.cell(row=row_idx,column=6,value=float(r.get("revenue",0))-float(r.get("cost",0)))
                ws3.cell(row=row_idx,column=7,value=float(r.get("margin_pct",0)))
                if row_idx%2==0:
                    for c in range(1,8): ws3.cell(row=row_idx,column=c).fill=alt_fill
            _auto_width(ws3)

            # Sheet 4: Inventory Rotation
            ws4 = wb.create_sheet("Rotación Inventario")
            _header_row(ws4,["Producto","Categoría","Stock","Vendido","Rotación","Días Agotamiento","Valor"])
            for row_idx,r in enumerate(rot,2):
                for col,k in enumerate(["nombre","categoria","stock","sold","rotation","days_to_stockout","valor"],1):
                    ws4.cell(row=row_idx,column=col,value=r.get(k,""))
                if row_idx%2==0:
                    for c in range(1,8): ws4.cell(row=row_idx,column=c).fill=alt_fill
            _auto_width(ws4)

            # Sheet 5: Loyal Clients
            ws5 = wb.create_sheet("Clientes Fieles")
            _header_row(ws5,["Cliente","Nivel","Puntos","Visitas","Total Compras","Margen Generado"])
            for row_idx,r in enumerate(top_c,2):
                ws5.cell(row=row_idx,column=1,value=r.get("nombre",""))
                ws5.cell(row=row_idx,column=2,value=r.get("nivel",""))
                ws5.cell(row=row_idx,column=3,value=int(r.get("puntos",0)))
                ws5.cell(row=row_idx,column=4,value=int(r.get("visitas",0)))
                ws5.cell(row=row_idx,column=5,value=float(r.get("total",0)))
                ws5.cell(row=row_idx,column=6,value=float(r.get("margen",0)))
                if row_idx%2==0:
                    for c in range(1,7): ws5.cell(row=row_idx,column=c).fill=alt_fill
            _auto_width(ws5)

            wb.save(output_path)
            total_rows = len(daily)+len(top_p)+len(rot)+len(top_c)
            self._log_export("CEO","EXCEL",output_path,exported_by,total_rows)
            return output_path
        except ImportError:
            raise RuntimeError("openpyxl not installed — cannot export Excel")
        except Exception as exc:
            logger.error("Excel export failed: %s", exc)
            raise
