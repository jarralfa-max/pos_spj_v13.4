
# core/services/export_service.py — SPJ POS v7  EXPORT CSV/EXCEL/PDF
from __future__ import annotations
import csv, io, json, logging, os
from datetime import datetime
from typing import List, Dict, Any, Optional
from core.db.connection import get_connection
logger = logging.getLogger("spj.export")

class ExportResult:
    def __init__(self, path:str, fmt:str, rows:int):
        self.path=path; self.format=fmt; self.rows=rows
    def __repr__(self): return f"<ExportResult {self.format} {self.rows}rows → {self.path}>"

class ExportService:
    """
    Servicio de exportación universal — CSV, Excel (xlsx via openpyxl), PDF (via reportlab).
    Si las librerías opcionales no están instaladas, ofrece fallback a CSV.
    """
    def __init__(self, conn=None):
        self.conn = conn or get_connection()

    # ── API pública ─────────────────────────────────────────────────────────
    def export(self, query:str, params:tuple=(), fmt:str="csv",
               filepath:str=None, title:str="Reporte", columns:list=None) -> ExportResult:
        rows = [dict(r) for r in self.conn.execute(query,params).fetchall()]
        if not rows: rows = []
        if columns is None and rows: columns = list(rows[0].keys())
        filepath = filepath or self._default_path(title, fmt)
        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
        if fmt=="csv":   return self._csv(rows, columns, filepath)
        if fmt=="xlsx":  return self._xlsx(rows, columns, filepath, title)
        if fmt=="pdf":   return self._pdf(rows, columns, filepath, title)
        raise ValueError(f"Formato no soportado: {fmt}")

    def export_ventas(self, desde:str=None, hasta:str=None, fmt:str="xlsx") -> ExportResult:
        sql="""SELECT v.id,v.folio,v.fecha,v.usuario,v.total,v.forma_pago,v.estado,
                      COUNT(d.id) as num_items
               FROM ventas v LEFT JOIN detalles_venta d ON d.venta_id=v.id
               WHERE v.estado!='cancelada'"""
        params=[]
        if desde: sql+=" AND v.fecha>=?"; params.append(desde)
        if hasta: sql+=" AND v.fecha<=?"; params.append(hasta)
        sql+=" GROUP BY v.id ORDER BY v.fecha DESC"
        return self.export(sql,tuple(params),fmt,title="Ventas")

    def export_inventario(self, fmt:str="xlsx") -> ExportResult:
        return self.export(
            "SELECT id,nombre,existencia,stock_minimo,precio,COALESCE(precio_compra,0) as costo,unidad,categoria FROM productos WHERE activo=1 ORDER BY nombre",
            fmt=fmt,title="Inventario")

    def export_clientes(self, fmt:str="xlsx") -> ExportResult:
        return self.export(
            "SELECT id,nombre,COALESCE(apellido,'') as apellido,COALESCE(telefono,'') as telefono,COALESCE(email,'') as email,COALESCE(puntos,0) as puntos FROM clientes WHERE activo=1 ORDER BY nombre",
            fmt=fmt,title="Clientes")

    def export_audit_logs(self, modulo:str=None, fmt:str="csv") -> ExportResult:
        sql="SELECT accion,modulo,entidad,usuario,sucursal_id,detalles,fecha FROM audit_logs"
        params=[]
        if modulo: sql+=" WHERE modulo=?"; params.append(modulo)
        sql+=" ORDER BY fecha DESC LIMIT 5000"
        return self.export(sql,tuple(params),fmt,title="AuditLog")

    def export_nomina(self, fmt:str="xlsx") -> ExportResult:
        return self.export("""
            SELECT p.nombre,COALESCE(p.apellido,'') as apellido,n.periodo_inicio,n.periodo_fin,
                   n.salario_base,n.deducciones,n.bonos,n.neto_pagar,n.estado
            FROM nomina_records n JOIN personal p ON p.id=n.personal_id
            ORDER BY n.fecha_registro DESC""",
            fmt=fmt, title="Nomina")

    # ── Backends ─────────────────────────────────────────────────────────────
    def _default_path(self, title:str, fmt:str) -> str:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe = title.replace(" ","_").lower()
        return os.path.join("exports", f"{safe}_{ts}.{fmt}")

    def _csv(self, rows:list, columns:list, filepath:str) -> ExportResult:
        with open(filepath,"w",newline="",encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
            w.writeheader(); w.writerows(rows)
        return ExportResult(filepath,"csv",len(rows))

    def _xlsx(self, rows:list, columns:list, filepath:str, title:str) -> ExportResult:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = title[:31]
            # Header row styling
            header_fill = PatternFill("solid",fgColor="1E3A5F")
            header_font = Font(bold=True,color="FFFFFF",size=10)
            thin = Side(style="thin",color="CCCCCC")
            border = Border(left=thin,right=thin,top=thin,bottom=thin)
            for col_i, col_nm in enumerate(columns,1):
                cell = ws.cell(1, col_i, col_nm.replace("_"," ").title())
                cell.font=header_font; cell.fill=header_fill
                cell.alignment=Alignment(horizontal="center",vertical="center")
                cell.border=border
            # Data rows
            alt_fill = PatternFill("solid",fgColor="F7F9FC")
            for row_i, row in enumerate(rows,2):
                fill = alt_fill if row_i%2==0 else PatternFill()
                for col_i, col_nm in enumerate(columns,1):
                    val = row.get(col_nm,"")
                    cell = ws.cell(row_i, col_i, val)
                    cell.fill=fill; cell.border=border
                    cell.alignment=Alignment(vertical="center")
            # Auto-width
            for col in ws.columns:
                max_len=max((len(str(c.value or "")) for c in col),default=8)
                ws.column_dimensions[col[0].column_letter].width=min(max_len+4,40)
            # Summary row
            ws.append([])
            ws.append([f"Total: {len(rows)} registros","Exportado:", datetime.now().strftime("%d/%m/%Y %H:%M")])
            wb.save(filepath)
            return ExportResult(filepath,"xlsx",len(rows))
        except ImportError:
            logger.warning("openpyxl no instalado — usando CSV fallback")
            csv_path = filepath.replace(".xlsx",".csv")
            return self._csv(rows,columns,csv_path)

    def _pdf(self, rows:list, columns:list, filepath:str, title:str) -> ExportResult:
        try:
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            doc = SimpleDocTemplate(filepath, pagesize=landscape(letter))
            styles = getSampleStyleSheet(); elements = []
            elements.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
            elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
            elements.append(Spacer(1,12))
            headers = [c.replace("_"," ").title() for c in columns]
            data = [headers]+[[str(row.get(c,"")) for c in columns] for row in rows[:500]]
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1E3A5F")),
                ("TEXTCOLOR",(0,0),(-1,0),colors.white),
                ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                ("FONTSIZE",(0,0),(-1,-1),8),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F7F9FC")]),
                ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#CCCCCC")),
                ("ALIGN",(0,0),(-1,-1),"CENTER"),
                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ]))
            elements.append(table)
            doc.build(elements)
            return ExportResult(filepath,"pdf",len(rows))
        except ImportError:
            logger.warning("reportlab no instalado — usando CSV fallback")
            csv_path = filepath.replace(".pdf",".csv")
            return self._csv(rows,columns,csv_path)

class ExportDialog:
    """Helper para mostrar diálogo de exportación en la UI."""
    @staticmethod
    def exportar(parent, conn, tipo:str="ventas"):
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QFormLayout, QComboBox, QDateEdit, QDialogButtonBox, QMessageBox, QFileDialog
        from PyQt5.QtCore import QDate
        dlg = QDialog(parent); dlg.setWindowTitle("Exportar"); dlg.setMinimumWidth(360)
        layout = QVBoxLayout(dlg); form = QFormLayout()
        combo_fmt = QComboBox(); combo_fmt.addItems(["Excel (.xlsx)","CSV (.csv)","PDF (.pdf)"])
        date_desde = QDateEdit(QDate.currentDate().addDays(-30)); date_desde.setCalendarPopup(True)
        date_hasta = QDateEdit(QDate.currentDate()); date_hasta.setCalendarPopup(True)
        form.addRow("Formato:", combo_fmt)
        if tipo=="ventas":
            form.addRow("Desde:", date_desde); form.addRow("Hasta:", date_hasta)
        layout.addLayout(form)
        btns = QDialogButtonBox(QDialogButtonBox.Ok|QDialogButtonBox.Cancel)
        btns.accepted.connect(dlg.accept); btns.rejected.connect(dlg.reject)
        layout.addWidget(btns)
        if dlg.exec_()!=QDialog.Accepted: return
        fmt_map={"Excel (.xlsx)":"xlsx","CSV (.csv)":"csv","PDF (.pdf)":"pdf"}
        fmt=fmt_map[combo_fmt.currentText()]
        svc=ExportService(conn)
        try:
            if tipo=="ventas":
                result=svc.export_ventas(date_desde.date().toString("yyyy-MM-dd"),date_hasta.date().toString("yyyy-MM-dd"),fmt)
            elif tipo=="inventario": result=svc.export_inventario(fmt)
            elif tipo=="clientes":   result=svc.export_clientes(fmt)
            else: result=svc.export(f"SELECT * FROM {tipo}",fmt=fmt,title=tipo)
            QMessageBox.information(parent,"Exportación completada",f"Archivo generado:\n{result.path}\n{result.rows} registros exportados.")
        except Exception as e:
            QMessageBox.critical(parent,"Error al exportar",str(e))
